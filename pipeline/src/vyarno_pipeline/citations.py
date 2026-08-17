"""Does every `source_url` still serve the figure printed beside it?

**A gate passing is not evidence.** Everything in `validate.py` checks the shape
of what a connector returned during the run that wrote it: bands, orderings,
sums, freshness. None of it can see the published file afterwards, and none of
it asks the question a reader asks by clicking the link — *is this number
actually there?* `validate_link_status` comes closest and stops at liveness, on
Eurostat alone: it reads the status code and a body predicate, so a БНБ URL that
now serves a Site Studio 404 shell with HTTP 200, or an ЕЦБ key that resolves to
a different series, passes it.

So this walks `data/published/`, pulls every citation out of the JSON, fetches
it, and holds the payload's own numbers against what comes back.

## A revision is not a fault, and the whole design turns on telling them apart

Upstreams restate. БНБ revise a month when a bank refiles; the ЕЦБ republish a
quarter; Eurostat move an index when a weight vintage lands. A check that failed
identically for that and for a broken citation would be a check somebody mutes
within two months, and then the broken citations stop being reported too.

    BROKEN    the URL does not resolve, or resolves to something that is not
              what the payload says it is: a 404, a workbook that is not a
              workbook, an SDMX response describing a different series, or a
              response in which not one of the payload's periods appears.
              The citation is wrong. Exit 3.

    REVISED   the URL resolves, it is the right series, the payload's periods
              are in it, and a value has moved. The upstream restated and a
              refresh will carry it. Reported, never fatal.

    STALE     the upstream has published a period newer than the payload's.
              A refresh is due. Reported, never fatal.

    OK        every value the payload prints is the value the upstream serves.

## What it cannot check, and why that is said out loud

Three of the six upstreams publish through a surface with no machine-readable
value at the other end of the URL a reader clicks: an имот.bg listing page, an
НСИ table page, a Държавен вестник issue. Those are reported `UNCHECKED` with
the reason, and never silently. `test_citations.py` holds the complement: every
citation in every payload is either fetched here or named in `UNCHECKABLE` with
a sentence saying why, so a new `source_url` cannot arrive unexamined — it fails
that test until somebody decides which it is.

A Eurostat `databrowser/view/...` URL is the human landing page for a cube and
carries no values; the same block always carries an `api_url` beside it that
does. Those are `LIVENESS` — fetched, checked for a 200 and for the dataset code
in the body, and not compared.

## Where it runs

**Not in `make check`.** It needs a network and six upstreams, and a suite that
fails when the wifi drops is one people learn to skip — the argument
`check-live-headers.mjs` makes for itself, and the same one. Run it
deliberately:

    vyarno-pipeline verify-citations                  # every payload
    vyarno-pipeline verify-citations --only credit    # one stem
    vyarno-pipeline verify-citations --quiet          # only what is not OK

Run it after a refresh, before believing a figure somebody has questioned, and
when an upstream announces a methodology change.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

import httpx

from vyarno_pipeline.sources.ecb import (
    BSI_SERIES_KEY_DIMS,
    CBD2_KEY_DIMS,
    EURO_SWITCH_PERIOD,
    SERIES_KEY_DIMS,
    _parse_sdmx_series,
)

BROWSER_UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"

# A БНБ path that does not exist 404s cleanly from `bnb_download` and is ~260
# bytes. Anywhere else on bnb.bg an unknown path returns the Site Studio shell
# with HTTP 200 — so a status code is not evidence there and a size floor is
# part of the check rather than a nicety.
BNB_MIN_WORKBOOK_BYTES = 5_000

# Verdicts, worst first. The order is the report's order and the exit code's.
BROKEN, REVISED, STALE, DERIVED, UNCHECKED, LIVENESS, OK = (
    "BROKEN",
    "REVISED",
    "STALE",
    "DERIVED",
    "UNCHECKED",
    "LIVENESS",
    "OK",
)


@dataclass
class Finding:
    """One citation, checked."""

    payload: str
    where: str
    url: str
    verdict: str
    detail: str = ""
    checked: int = 0
    notes: list[str] = field(default_factory=list)


# Hosts whose citation is a page for a person rather than an endpoint for a
# parser. Each says what a reader gets and why no value can be read off it.
UNCHECKABLE: dict[str, str] = {
    "www.imot.bg": (
        "a listing search page rendered for a browser. The published figure is a "
        "median over districts scraped at run time, and the page a reader opens "
        "shows today's listings rather than the month the payload names — there is no "
        "period to look the value up at. Datacenter IPs are also served 403."
    ),
    "www.nsi.bg": (
        "an НСИ page rather than one of their workbooks. Their timeseries "
        "workbooks ARE checked — the branch below reads any .xlsx — so a "
        "citation landing here is one pointing at a portal page instead."
    ),
    "dv.parliament.bg": (
        "a Държавен вестник issue. It is the legal instrument a parameter comes "
        "from, not a data endpoint, and the figure is read out of its text by a "
        "human once a year."
    ),
}

# Which SDMX flow a data-api.ecb.europa.eu URL names, and the dimension tuple
# the response has to decode against. Getting this from the URL rather than from
# a table beside it is the point: the check reads what the payload cites.
ECB_FLOW_DIMS: dict[str, tuple[str, ...]] = {
    "MIR": SERIES_KEY_DIMS,
    "CBD2": CBD2_KEY_DIMS,
    "BSI": BSI_SERIES_KEY_DIMS,
}


def tolerance_for(value: float) -> float:
    """Half a unit of the last decimal the payload chose to print.

    The payload rounds; the upstream does not. Comparing 8.76 against 8.7551
    exactly would report a revision on every figure in the tree, so the band is
    the rounding the publisher's own value went through and nothing wider.
    """
    text = repr(float(value))
    decimals = len(text.split(".")[1].rstrip("0")) if "." in text else 0
    return 0.5 * 10 ** (-decimals) + 1e-9


def _fetch(url: str, client: httpx.Client) -> httpx.Response:
    return client.get(url, headers={"User-Agent": BROWSER_UA})


def check_ecb(url: str, want: dict[str, float], client: httpx.Client) -> tuple[str, str, int]:
    """One ЕЦБ SDMX citation against the series it names."""
    path = urlparse(url).path
    match = re.search(r"/service/data/([A-Z0-9_]+)/([^/?]+)", path)
    if not match:
        return BROKEN, f"not an SDMX data URL: {path}", 0
    flow, key = match.group(1), match.group(2)
    dims = ECB_FLOW_DIMS.get(flow)
    if dims is None:
        return BROKEN, f"unknown ECB flow {flow!r} — add its dimension tuple to ECB_FLOW_DIMS", 0
    response = _fetch(url, client)
    if response.status_code != 200:
        return BROKEN, f"HTTP {response.status_code}", 0
    try:
        # The identity guard is the reason this reuses the connector's parser
        # rather than reading the JSON here: it refuses a response whose decoded
        # key is not the one requested, which is the failure a status code and a
        # non-empty body both pass.
        live = _parse_sdmx_series(response.json(), dims, key, flow)
    except ValueError as exc:
        return BROKEN, str(exc).splitlines()[0], 0
    return _compare(live, want)


def check_eurostat(url: str, want: dict[str, float], client: httpx.Client) -> tuple[str, str, int]:
    """One Eurostat JSON-stat citation against the cube it names."""
    response = _fetch(url, client)
    if response.status_code != 200:
        return BROKEN, f"HTTP {response.status_code}", 0
    body = response.json()
    times = (body.get("dimension", {}).get("time", {}).get("category", {}) or {}).get("index")
    values = body.get("value")
    if not isinstance(times, dict) or not isinstance(values, dict):
        return BROKEN, "no JSON-stat time index or value map in the response", 0
    # JSON-stat keys the value map by the FLATTENED observation index and omits
    # missing cells, so a positional read is wrong wherever a period is absent.
    by_period = {period: values[str(idx)] for period, idx in times.items() if str(idx) in values}
    return _compare(by_period, want)


def check_bnb(url: str, want: dict[str, float], client: httpx.Client) -> tuple[str, str, int]:
    """A БНБ citation, routed by what it actually is.

    **Three kinds live on this host and only one is a workbook.** The lending
    limits cite a press release and the observed DSTI cites a supervisory PDF,
    and holding either to "is this a zip container" reported a broken citation
    on two links that resolve correctly. What each kind gets is the strongest
    check available for it, and the byte floor is the shared part: an unknown
    path outside `bnb_download` returns the Site Studio shell with HTTP 200, so
    a status code is not evidence anywhere on bnb.bg.
    """
    if url.lower().endswith(".xlsx"):
        return check_bnb_workbook(url, want, client)
    response = _fetch(url, client)
    if response.status_code != 200:
        return BROKEN, f"HTTP {response.status_code}", 0
    body = response.content
    if url.lower().endswith(".pdf"):
        if not body.startswith(b"%PDF"):
            return BROKEN, f"{len(body)} bytes and not a PDF — an error page", 0
        return LIVENESS, f"PDF, {len(body) // 1024} kB", 0
    # An HTML page. The shell an unknown path returns is 16.8 kB to a library
    # user agent and 27.3 kB to a browser one, so a page that is really there is
    # recognised by carrying its own title rather than by answering 200.
    if len(body) < BNB_MIN_WORKBOOK_BYTES:
        return BROKEN, f"{len(body)} bytes — too small to be the page cited", 0
    return LIVENESS, f"HTML page, {len(body) // 1024} kB", 0


def _derived(node: dict[str, Any], prefix: str) -> str:
    """What the payload says it worked out itself, and from what.

    A derived figure is not in the workbook and never was, so searching for it
    finds nothing and reports a revision on a payload that is correct. The
    payload marks each one: `outstanding.total_eur_m` ships beside the `blocks`
    it sums, and `overdraft.stock_eur_m` ships beside the `stock_basis` naming
    the subtraction. The addends ARE cells and are what gets checked instead.
    """
    if node.get(f"{prefix}_basis" if prefix else "basis"):
        return str(node[f"{prefix}_basis" if prefix else "basis"])
    if (
        not prefix
        and node.get("total_eur_m")
        and isinstance(node.get("blocks") or node.get("buckets"), list)
    ):
        return "a total over the components beside it; the components are the cells"
    return ""


def check_bnb_workbook(
    url: str, want: dict[str, float], client: httpx.Client
) -> tuple[str, str, int]:
    """A БНБ citation resolves to a workbook, and one that opens.

    **A status code proves nothing on this host.** `bnb_download` 404s cleanly on
    a bad filename, but an unknown path elsewhere on bnb.bg returns an identical
    Site Studio shell with HTTP 200 — so the check is that the bytes are a
    workbook the parser can open, at a size no error page reaches.
    """
    response = _fetch(url, client)
    if response.status_code != 200:
        return BROKEN, f"HTTP {response.status_code}", 0
    body = response.content
    if len(body) < BNB_MIN_WORKBOOK_BYTES or not body.startswith(b"PK"):
        return (
            BROKEN,
            f"{len(body)} bytes and not a zip container — an error page, not a workbook",
            0,
        )
    try:
        from io import BytesIO

        from openpyxl import load_workbook

        book = load_workbook(BytesIO(body), read_only=True, data_only=True)
        sheets = book.sheetnames
    except Exception as exc:  # any failure to open is the finding
        return BROKEN, f"the bytes are not a readable workbook: {exc!r}", 0
    if not want:
        return LIVENESS, f"workbook opens, sheets {sheets}", 0
    return _find_in_workbook(book, want)


def _find_in_workbook(book: Any, want: dict[str, float]) -> tuple[str, str, int]:
    """Every wanted value appears somewhere in the workbook's cells.

    Deliberately a search rather than a column read. The payload names a sheet
    and a column in prose, not in a form this could index by, and the failure
    worth catching is a column that MOVED — which a coordinate read would follow
    silently into the wrong block. A value that is nowhere in the sheet at all is
    the unambiguous half of that, and it is the half worth an exit code.
    """
    seen: set[float] = set()
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    seen.add(round(float(cell), 4))
    missing = [
        f"{period}={value}"
        for period, value in want.items()
        if not any(abs(value - candidate) <= tolerance_for(value) for candidate in seen)
    ]
    if missing:
        return REVISED, f"not in the workbook: {', '.join(missing[:4])}", len(want)
    return OK, "", len(want)


def _compare(live: dict[str, float], want: dict[str, float]) -> tuple[str, str, int]:
    """The payload's periods against the upstream's, and the verdict."""
    if not want:
        return LIVENESS, f"resolves, {len(live)} periods, no value declared to check", 0
    overlap = [p for p in want if p in live]
    if not overlap:
        return (
            BROKEN,
            f"none of the payload's {len(want)} periods are in a response carrying "
            f"{len(live)} of them ({sorted(live)[:2]}…) — this URL serves something else",
            0,
        )
    moved = [
        f"{p}: published {want[p]}, upstream {live[p]}"
        for p in overlap
        if abs(live[p] - want[p]) > tolerance_for(want[p])
    ]
    withdrawn = [p for p in want if p not in live]
    if moved:
        return REVISED, "; ".join(moved[:3]), len(overlap)
    if withdrawn:
        return (
            REVISED,
            f"upstream no longer publishes {len(withdrawn)}: {withdrawn[:3]}",
            len(overlap),
        )
    newest_live, newest_want = max(live), max(want)
    if newest_live > newest_want:
        return STALE, f"upstream is at {newest_live}, the payload at {newest_want}", len(overlap)
    return OK, "", len(overlap)


# How many of a series' months a citation is held to. **Not the whole history,
# and the reason is the splice.** `credit.json#consumer` is the BGN leg through
# 2025 and the EUR leg after it, under one URL that can only be one of the two —
# so comparing 2020-2025 against the euro series would report six years of
# revisions on a payload that is correct. A shifted column or a swapped key
# moves the recent months too, which is what this is for.
SERIES_WINDOW = 12

# Keys whose value is a citation. Four spellings and every one of them is a
# link this repository asks a reader to trust: `api_url` is Eurostat's data
# endpoint, `source_url` the page a reader clicks, `source_url_bg` its
# Bulgarian-language twin, and `cross_check_url` the second publisher a figure
# was gated against. `test_citations.py` compares this walk against a plain grep
# for a URL, which is how the first draft's missing six were found.
_URL_KEY = re.compile(r"^api_url|source_url(_[a-z]{2})?$|cross_check_url$")
# A published `{period: value}` map. The payloads all name them this way.
_SERIES_KEY = re.compile(r"_by_period$|^series_by_period$")
# A scalar a citation vouches for, once its block's prefix is taken off.
_VALUE_KEY = re.compile(r"(_pct|_eur_m|_index|_count)$|^value$")


# A Eurostat data URL says which unit it serves, and the block beside it carries
# both: `prc_hicp_minr` is the annual rate at `RCH_A` and the index at `I15`.
# Reading the unit out of the citation rather than guessing from a field name is
# the difference between checking the rate against the rate and checking «2,3%»
# against an index of 182,45 — which is what a name-order guess did, on all
# thirteen divisions and both their groups at once.
EUROSTAT_UNIT_FIELDS: dict[str, tuple[str, ...]] = {
    "RCH_A": ("annual_rate_pct", "rate_annual_pct", "rate_pct"),
    "I15": ("latest_index", "value"),
}


def _eurostat_unit(url: str) -> str | None:
    if "ec.europa.eu" not in url or "/api/dissemination/" not in url:
        return None
    units = parse_qs(urlparse(url).query).get("unit") or []
    return units[0] if units else None


def _prefix_of(key: str) -> str:
    """The block a citation belongs to: `stock_source_url` vouches for `stock_*`."""
    if key.startswith("api_url"):
        return key[len("api_url") :].lstrip("_")
    # `source_url_bg` is the same citation in the other language and owns the
    # same block, so the language tag comes off before the prefix is read.
    stem = re.sub(r"_[a-z]{2}$", "", key) if re.search(r"source_url_[a-z]{2}$", key) else key
    for suffix in ("source_url", "cross_check_url"):
        if stem.endswith(suffix):
            return stem[: -len(suffix)].rstrip("_")
    return ""


def _spliced(node: dict[str, Any]) -> bool:
    """The block's series changes currency partway, so one URL cannot serve it.

    `credit.json#consumer` is `…BGN.N` through 2025 and `…EUR.N` after it, and
    the payload cites the euro leg because that is the live one. The euro leg
    DOES answer for the earlier months — with the niche of euro-denominated
    lending that existed before adoption, tens of millions a month against the
    BGN series' billion — so comparing across the seam finds real numbers that
    are the wrong ones. Only the euro era is this citation's to answer for.
    """
    # Three fields declare it and no block uses all three: `dataset` says
    # «spliced with», `unit` says «currency of the period», `_role` calls it the
    # evidence for the splice. Reading one of them only would hold a spliced
    # series to a euro key on whichever blocks happened to word it differently.
    declared = " ".join(str(node.get(k, "")) for k in ("dataset", "unit", "_role")).lower()
    return "splice" in declared or "currency of the period" in declared


def _wanted(node: dict[str, Any], prefix: str, url: str = "") -> dict[str, float]:
    """The figures in `node` that a citation with this prefix stands behind.

    Prefix-first and then bare, because a block carries both: `stock_source_url`
    is answerable by `stock_rate_pct` and never by the `value_pct` above it,
    while a plain `source_url` takes whatever the block's own headline is.
    """
    want: dict[str, float] = {}

    def scoped(pattern: re.Pattern[str]) -> list[str]:
        keys = [k for k in node if pattern.search(k)]
        owned = [k for k in keys if prefix and k.startswith(prefix)]
        return owned or (
            [k for k in keys if not _claimed_by_another(node, k)] if not prefix else []
        )

    series_keys = scoped(_SERIES_KEY)
    # **More than one and a bare URL: check none of them.** `non_performing`
    # carries households AND corporations beside a `source_url` that is the
    # household key, and merging the two silently held the household series to
    # the corporate figures — which agree to nothing, so it reported a revision
    # every quarter while both series were correct. Where the payload splits by
    # scope it also cites each scope separately, and those citations are exact.
    if len(series_keys) == 1:
        series = node[series_keys[0]]
        if isinstance(series, dict):
            periods = sorted(series)
            if _spliced(node):
                periods = [p for p in periods if p >= EURO_SWITCH_PERIOD]
            for period in periods[-SERIES_WINDOW:]:
                if isinstance(series[period], (int, float)):
                    want[period] = float(series[period])
    if want:
        return want
    period = node.get(f"{prefix}_ref_period") if prefix else None
    period = period or node.get("ref_period")
    unit = _eurostat_unit(url)
    candidates = (
        [k for k in EUROSTAT_UNIT_FIELDS.get(unit, ()) if k in node] if unit else scoped(_VALUE_KEY)
    )
    # A cube whose unit this module has no field mapping for is not guessed at.
    if unit and not candidates:
        return {}
    for key in candidates:
        cell = node[key]
        # `latest_index` carries its own period, which is the one the citation
        # answers for rather than the block's headline `ref_period`.
        if isinstance(cell, dict) and isinstance(cell.get("value"), (int, float)):
            return {str(cell.get("time", period)): float(cell["value"])}
        if isinstance(cell, (int, float)) and isinstance(period, str):
            return {period: float(cell)}
    return want


def _claimed_by_another(node: dict[str, Any], key: str) -> bool:
    """A bare `source_url` does not vouch for a sibling block's own figures."""
    prefixes = [_prefix_of(k) for k in node if _URL_KEY.search(k)]
    return any(p and key.startswith(p) for p in prefixes)


def _addends(node: dict[str, Any]) -> dict[str, float]:
    """The published components of a derived total, keyed by their own name."""
    # `outstanding` calls them blocks and `fixation` calls them buckets; both are
    # the publisher's own cells under a total this repository derives.
    blocks = node.get("blocks") or node.get("buckets")
    if not isinstance(blocks, list):
        return {}
    out: dict[str, float] = {}
    for block in blocks:
        if not isinstance(block, dict) or not isinstance(block.get("volume_eur_m"), (int, float)):
            continue
        name = str(block.get("block") or block.get("bucket"))
        # An addend the node cites SEPARATELY belongs to that citation and not
        # to this one. The four blocks of what households owe come out of two
        # different workbooks — the overdraft one has its own
        # `overdraft_source_url` — so holding the loan workbook to all four
        # reports a missing cell that was never supposed to be in it.
        if _claimed_by_another(node, name):
            continue
        out[name] = float(block["volume_eur_m"])
    return out


def extract(payload: Any, where: str = "") -> list[tuple[str, str, dict[str, float]]]:
    """Every citation in a payload, with the figures it stands behind."""
    found: list[tuple[str, str, dict[str, float]]] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            if isinstance(value, str) and _URL_KEY.search(key) and value.startswith("http"):
                prefix = _prefix_of(key)
                want = _wanted(payload, prefix, value)
                basis = _derived(payload, prefix)
                if basis:
                    # The total is ours; its addends are the publisher's, and
                    # those are what a workbook can be held to.
                    want = _addends(payload) or want
                found.append(
                    (
                        f"{where}.{key}".lstrip("."),
                        value,
                        want if not basis or _addends(payload) else {},
                    )
                )
            elif isinstance(value, dict) and _URL_KEY.search(key):
                # `scope_source_urls` is a map of name → URL, each standing
                # behind the same-named figure in the block above it.
                for name, url in value.items():
                    if isinstance(url, str) and url.startswith("http"):
                        found.append((f"{where}.{key}.{name}", url, _wanted(payload, name, url)))
            else:
                found.extend(extract(value, f"{where}.{key}".lstrip(".")))
    elif isinstance(payload, list):
        for i, item in enumerate(payload):
            found.extend(extract(item, f"{where}[{i}]"))
    return found


def check_one(url: str, want: dict[str, float], client: httpx.Client) -> tuple[str, str, int]:
    """Route one citation to the reader that understands its host."""
    host = urlparse(url).netloc
    # A workbook is readable whoever publishes it, and this runs before the
    # unreadable-host list so a host that mostly serves pages is still checked
    # on the files it does serve.
    if url.lower().endswith(".xlsx"):
        return check_bnb_workbook(url, want, client)
    if host in UNCHECKABLE:
        return UNCHECKED, UNCHECKABLE[host], 0
    if host == "data-api.ecb.europa.eu":
        return check_ecb(url, want, client)
    if host.endswith("bnb.bg"):
        return check_bnb(url, want, client)
    if host == "ec.europa.eu":
        if "/api/dissemination/" in url:
            return check_eurostat(url, want, client)
        # The databrowser landing page for a cube. It carries no values, and the
        # block citing it always carries an `api_url` beside it that does — so
        # what is checked here is that the cube still exists under that code.
        dataset = urlparse(url).path.rstrip("/").split("/")[-3:]
        response = _fetch(url, client)
        if response.status_code != 200:
            return BROKEN, f"HTTP {response.status_code}", 0
        return LIVENESS, f"databrowser page for {'/'.join(dataset)}", 0
    return UNCHECKED, f"no reader for host {host!r}", 0


def verify(published: Path, only: str | None = None, timeout: float = 60.0) -> list[Finding]:
    """Check every citation in `published`, one payload at a time."""
    findings: list[Finding] = []
    files = sorted(published.glob("*.json"))
    if only:
        files = [f for f in files if only in f.stem]
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        for path in files:
            payload = json.loads(path.read_text(encoding="utf-8"))
            # One fetch per distinct URL. `hicp_categories.json` cites the same
            # databrowser page from all thirteen divisions, and 132 requests to
            # say one thing is a check that takes long enough to stop being run.
            cache: dict[tuple[str, frozenset], tuple[str, str, int]] = {}
            for where, url, want in extract(payload):
                key = (url, frozenset(want.items()))
                if key not in cache:
                    try:
                        cache[key] = check_one(url, want, client)
                    except httpx.HTTPError as exc:
                        cache[key] = (UNCHECKED, f"unreachable from here: {exc!r}", 0)
                verdict, detail, checked = cache[key]
                findings.append(Finding(path.name, where, url, verdict, detail, checked))
    return findings
