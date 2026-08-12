"""НСИ XLSX connectors — the regional and the by-sector average gross wage.

    https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.2_EUR_EN.xlsx
    https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.2_EUR.xlsx
    https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.1_EUR_EN.xlsx
    https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.1_EUR.xlsx

Two sibling tables from one publisher, in EUR: `1.1.2.2` by region and `1.1.2.1`
by economic activity. Both are read for НСИ's own **published quarterly
averages** — never a month, and never a figure this code computed.

The regional average gross wage table (BG statistical regions and districts),
in EUR. We read the `{year}trimes` sheets — НСИ's own **published quarterly
averages** by region — for **all 28 области**, which is every district row the
sheet carries. Its own title says `BY STATISTICAL REGIONS AND DISTRICTS`, so the
whole country was always inside the file a single-row read was opening.

**Both language editions, for the same reason the by-sector table reads both.**
The область names are half of what this payload is for — a picker has to print
them — and translating НСИ's English ourselves is how «София(столица)» becomes
something they never wrote. So each label is the one НСИ printed in that
language, and the two editions pin each other cell for cell.

**The two editions do not have the same row layout.** The Bulgarian one carries
an extra «Общо» row above the unit marker, so its quarter headers sit one row
lower, and its sheets are named `'2026 trimes '` where the English ones are
`2026trimes`. Neither the header row nor the sheet name may therefore be an
index or a literal: the header row is found by being the first row that parses
as four quarters, and the sheets by a regex tolerant of the spacing.

**Why the quarter and not the month.** BG wages have a strong seasonal shape and
March is the annual peak, as Q1 bonus and 13th-salary payments land: the
March-over-February step runs +7% to +13% and has grown every year. Anchoring
the salary distribution to a single month scales every rung by whatever seasonal
factor the last refreshed month happened to carry.

**Why НСИ's published quarter and not an average of their months.** The
workbook carries both, and the published quarter is the one to take. It is
authoritative rather than persuasive: for Q1 2026 НСИ publish 1915, where the
mean of their three rounded monthly cells gives 1914.7 — a rounding error with
nothing to buy it. And it keeps every figure in the payload one НСИ published,
which is what §2.1.1 of their licence needs, since it forbids distributing
производни и сборни произведения (docs/legal.md §НСИ).

**The Q4 trap.** Every `{year}trimes` sheet publishes Q4 twice: `IV` and
`IV incl.annual bonuses`. The two diverge by 6-8% (2025: 1859 against 2009)
because the second folds in the 13th salary. Taking the wrong one would step the
whole ladder up every fourth quarter and back down in Q1. We take `IV`, and
`_quarter_columns` refuses any column whose header mentions a bonus.

**The header trap.** The quarter headers mix alphabets — Q1 and Q2 are Cyrillic
І (U+0406), Q3 and Q4 are Latin I and V. `_roman_quarter` folds the Cyrillic
form onto the Latin one rather than matching literals, so the parse survives НСИ
normalising the encoding either way.

**Why EUR, not BGN.** The same table exists in BGN but lags by one quarterly
release cycle. Every other number in the SPA is EUR, so publishing in EUR avoids
a unit mismatch.

**Why the XLSX, not the HTML landing page.** The page uses rowspan/colspan
headers that roll forward every quarter and break naive parsers. The XLSX has a
stable schema — the same column layout for every year sheet.

**The regression guard, and what it can assert now that every row is read.**
A single-row read could only compare that row against a neighbour, so it
asserted `-Sofia cap.` above `-Sofia` (the province) and there was no universal
"region X beats region Y" to generalise from. Reading the whole table affords a
stronger guard in three parts, and all three have to hold:

1. **Every one of the 28 области named in `regions.py` is present**, in both
   editions. A renamed row fails here rather than going missing from a picker.
2. **No district row is present that `regions.py` does not name.** Without this
   the guard is one-directional: НСИ splitting an област would add a row nobody
   notices, and the payload would carry 28 of 29.
3. **Sofia-city is the maximum.** It is not a taste — it is the highest-wage
   region in Bulgaria by a wide margin (1915 against a next-highest 1304 at
   2026-Q1), and a selector that drifted onto the statistical-region rows above
   the districts, or off by one row, breaks it.

What the three catch together that the old one could not: an off-by-one that
shifts every reading by one область keeps `cap > province` true while putting
Varna's wage under Dobrich's name.

THE BY-SECTOR TABLE (`Labour_1.1.2.1`)
--------------------------------------

19 NACE Rev 2 sections plus `Total`, on the same quarterly rhythm. The three
traps above all apply to it unchanged; three more are specific to it.

**Both language editions are read, because the labels are the product.** НСИ
publish the identical table twice — `_EUR_EN.xlsx` with English section names,
`_EUR.xlsx` with Bulgarian ones — and the site needs both. Translating НСИ's
English ourselves would be the whole hazard of this feature in one step: their
Bulgarian name for section J is «Създаване и разпространение на информация и
творчески продукти; далекосъобщения», which nobody mistakes for «ИТ», while a
translation of "Information and communication" invites exactly that reading. So
each label is the one НСИ printed in that language.

**The two editions pin each other.** They carry the same figures, so the rows
are paired by position and every paired cell must be equal. That is what makes
positional pairing safe: if НСИ reorder or insert a row in one edition, the
values stop matching and the parse raises instead of shipping section J's wage
under section K's name.

**Each sheet stacks FOUR blocks, and two of them are titles.** Monthly by
activity, monthly by ownership, quarterly by activity, quarterly by ownership —
and `Total` appears in column 0 of four rows, two of which are section titles
with no data at all. A label scan finds a title row before either data row.
Every block here is therefore bounded from the header row it was found by, and
terminated by the first blank label, so no read can cross into the next block.
"""

from __future__ import annotations

import io
import re
from typing import Any

import httpx
import openpyxl

from vyarno_pipeline.regions import REGIONS, SOFIA_CITY_CODE

# Canonical URLs. If НСИ moves either,
# `tests/test_nsi.py::test_connector_url_is_nsi_timeseries_xlsx` fails before a
# mis-extracted number can reach production.
SOURCE_URL = (
    "https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.2_EUR_EN.xlsx"
)
# The Bulgarian edition of the same table carries no language suffix, exactly as
# the by-sector pair below does.
SOURCE_URL_BG = (
    "https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.2_EUR.xlsx"
)

# One sheet per year of published quarterly averages. The English edition names
# them "2026trimes"; the Bulgarian one "2026 trimes " — same table, different
# whitespace — so the year is matched rather than the string stripped. The
# workbook also carries a combined monthly sheet ("2019-2026"), which the
# four-digit anchor excludes: the per-year sheets are the ones that label their
# own year and carry the preliminary marker in their title.
_QUARTERLY_SHEET_RE = re.compile(r"^(\d{4})\s*trimes\s*$", re.IGNORECASE)

# How far into a sheet the quarter headers may sit before we give up. The
# English edition puts them on row 4 and the Bulgarian one on row 5, because it
# carries an extra «Общо» row; the search below is what absorbs that difference
# and any further row НСИ inserts above the table.
_MAX_HEADER_SEARCH_ROW = 12

# A district row is prefixed with a hyphen in НСИ's own layout — that is how
# their sheet distinguishes an област from the statistical region heading it
# sits under, and it is what lets the parse notice a district row this project
# does not know about.
_DISTRICT_PREFIX = "-"

# Q4 is published twice and the second column includes the 13th salary. Taking
# it would step the ladder up every fourth quarter — see the module docstring.
#
# Both spellings, because the by-sector table is read in both language editions
# and the Bulgarian one heads that column «IV вкл.годишни премии».
#
# The exact match at the end of `_roman_quarter` rejects both spellings on its
# own today, so these markers change no behaviour against the current headers.
# They are what holds if that exact match is ever relaxed to tolerate a trailing
# note in the header — a reasonable-looking change, since it would let the parse
# survive НСИ appending a footnote marker to "IV". Loosen it with only the
# English marker present and the Bulgarian file starts reading Q4 as the
# bonus-inclusive column, which is 12% higher and looks like a good year.
_BONUS_MARKERS = ("bonus", "премии")

_ROMAN_TO_QUARTER = {"I": 1, "II": 2, "III": 3, "IV": 4}

# The quarter headers mix alphabets: Cyrillic І (U+0406) for Q1/Q2, Latin I and
# V for Q3/Q4. Fold the Cyrillic letters onto their Latin twins so the parse
# does not depend on which НСИ used this year.
_CYRILLIC_TO_LATIN = str.maketrans({"\u0406": "I", "\u0456": "I", "\u0430": "a"})


def _roman_quarter(header: object) -> int | None:
    """Quarter number from a header cell, or None if it is not one.

    Returns None for the "IV incl.annual bonuses" column, which is a different
    quantity wearing a quarter's name.
    """
    if header is None:
        return None
    text = str(header).strip().translate(_CYRILLIC_TO_LATIN)
    lowered = text.lower()
    if any(marker in lowered for marker in _BONUS_MARKERS):
        return None
    return _ROMAN_TO_QUARTER.get(text.upper())


def _quarter_columns(header_row: list) -> dict[int, int]:
    """Map sheet column index -> quarter number, for the four plain quarters."""
    cols = {
        i: q
        for i, cell in enumerate(header_row)
        if i > 0 and (q := _roman_quarter(cell)) is not None
    }
    if len(cols) != 4 or sorted(cols.values()) != [1, 2, 3, 4]:
        raise ValueError(
            f"Expected four quarter columns I..IV in the header row, got {cols}. "
            f"НСИ may have restructured the quarterly sheets, or renamed the "
            f"annual-bonus column so that it now parses as a quarter."
        )
    return cols


def _quarter_header_row(rows: list, sheet_title: str) -> int:
    """Index of the row carrying the four quarter headers.

    Searched rather than indexed, because the two language editions differ: the
    Bulgarian one has an extra «Общо» row above the unit marker, so its headers
    sit one row lower than the English one's. An index that is right for one
    file reads a blank row in the other and reports "no quarter columns", which
    is a true error message about the wrong thing.

    The first row that parses as exactly I..IV wins. Nothing above the table can
    match: `_quarter_columns` needs four distinct Roman numerals in columns
    beyond the first, and a title row is one string in column 0.
    """
    for i, row in enumerate(rows[:_MAX_HEADER_SEARCH_ROW]):
        try:
            _quarter_columns(row)
        except ValueError:
            continue
        return i
    raise ValueError(
        f"No quarter header row (I..IV) in the first {_MAX_HEADER_SEARCH_ROW} "
        f"rows of sheet {sheet_title!r}. НСИ may have restructured the quarterly "
        f"sheets, or renamed the annual-bonus column so that it now parses as a "
        f"quarter."
    )


def _parse_quarterly_sheet(ws, labels: dict[str, str]) -> dict[str, Any]:
    """One `{year}trimes` sheet -> every known област's quarterly row.

    `labels` maps НСИ's own row label, in this edition's language, to the
    region code. Returns `{"year", "is_preliminary", "by_code": {code: {q:
    eur}}, "unknown_districts": [label, ...]}`. Quarters with no value yet are
    absent, which is how the current year carries only the quarters published
    so far.

    `unknown_districts` is collected rather than raised on, so the caller can
    name every one of them at once instead of a run failing on the first.
    """
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    match = _QUARTERLY_SHEET_RE.match(str(ws.title).strip())
    if match is None:
        raise ValueError(f"Sheet {ws.title!r} is not a per-year quarterly sheet.")
    year = int(match.group(1))

    header_row = _quarter_header_row(rows, str(ws.title))
    cols = _quarter_columns(rows[header_row])

    # The title carries the preliminary marker for the whole year, e.g.
    # "... IN 2026*" / "... ПРЕЗ 2026 ГОДИНА*". НСИ drop the star when the year
    # is final. It is the first non-empty label above the header row, which is
    # what keeps this off a fixed index in a file whose two editions differ by
    # one row.
    title = next(
        (str(r[0]) for r in rows[:header_row] if r and r[0] is not None and str(r[0]).strip()),
        "",
    )
    is_preliminary = title.rstrip().endswith("*")

    by_code: dict[str, dict[int, float]] = {}
    unknown: list[str] = []
    for row in rows[header_row + 1 :]:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if not label.startswith(_DISTRICT_PREFIX):
            # A statistical-region heading, the unit marker, or the footnote.
            # None of them is an област and none is published.
            continue
        code = labels.get(label)
        if code is None:
            unknown.append(label)
            continue
        by_code[code] = {
            q: float(row[col])
            for col, q in cols.items()
            if col < len(row) and isinstance(row[col], (int, float))
        }

    return {
        "year": year,
        "is_preliminary": is_preliminary,
        "by_code": by_code,
        "unknown_districts": unknown,
    }


def _latest_published_quarter(quarters: dict[str, float]) -> str | None:
    """The most recent "YYYY-Qn" key present, or None if there are none.

    "YYYY-Qn" sorts lexicographically as chronologically for any 4-digit year,
    which is why the key is built that way rather than as a tuple.
    """
    return max(quarters) if quarters else None


def _get_xlsx(url: str, timeout: float = 30.0) -> bytes:
    """Single GET with sane defaults. Raises on non-2xx."""
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        r = client.get(url)
        r.raise_for_status()
        return r.content


def _parse_region_edition(raw: bytes, labels: dict[str, str], url: str) -> dict[str, Any]:
    """One language edition -> `{code: {"YYYY-Qn": eur}}` plus the marker flags.

    Raises on a district row this project does not name. Doing that here rather
    than in the caller means the error says which file it was reading, and both
    editions are checked rather than only the one the names are taken from.
    """
    import io

    # NOT read_only: the quarterly sheets are small and `iter_rows` over a
    # read-only worksheet reports dimensions unreliably on these files.
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sheets = [n for n in wb.sheetnames if _QUARTERLY_SHEET_RE.match(str(n).strip())]
    if not sheets:
        raise ValueError(
            f"No per-year quarterly sheet matching {_QUARTERLY_SHEET_RE.pattern} "
            f"found in {url}. Available sheets: {wb.sheetnames}"
        )

    series: dict[str, dict[str, float]] = {}
    prelim_by_period: dict[str, bool] = {}
    unknown: set[str] = set()
    for name in sorted(sheets):
        parsed = _parse_quarterly_sheet(wb[name], labels)
        unknown.update(parsed["unknown_districts"])
        year = parsed["year"]
        for code, values in parsed["by_code"].items():
            for q, v in values.items():
                series.setdefault(code, {})[f"{year}-Q{q}"] = v
                prelim_by_period[f"{year}-Q{q}"] = parsed["is_preliminary"]

    if unknown:
        raise ValueError(
            f"{url} carries district row(s) {sorted(unknown)!r} that "
            f"`regions.py#REGIONS` does not name. НСИ have added or renamed an "
            f"област. Refusing to publish a table that is missing one: add the "
            f"row to that table, with имот.bg's slug for it or None."
        )

    missing = [r.code for r in REGIONS if r.code not in series]
    if missing:
        raise ValueError(
            f"{url} is missing {len(missing)} of the {len(REGIONS)} области "
            f"`regions.py#REGIONS` names: {missing!r}. НСИ likely renamed a row "
            f"— the label in that table has to be their exact string."
        )
    return {"series": series, "prelim_by_period": prelim_by_period}


def fetch_region_salaries_eu(
    url_en: str = SOURCE_URL,
    url_bg: str = SOURCE_URL_BG,
) -> dict[str, Any]:
    """Fetch НСИ's regional wage table in both editions — every област.

    Returns:
        {
            "ref_period": "YYYY-Qn",        # latest quarter EVERY област carries
            "is_preliminary": bool,
            "regions": [
                {"code": str, "en_name": str, "bg_name": str,
                 "value_eur": float, "series_by_period": {"YYYY-Qn": float, ...}},
                ...                          # `regions.py#REGIONS` order
            ],
        }

    Every value returned is a cell НСИ published. Nothing here is averaged,
    rebased or interpolated — see the module docstring on why that matters.

    Raises:
        httpx.HTTPError — network failure
        ValueError      — no quarterly sheet, structure changed, the two
                          editions disagree, or a regression guard tripped
    """
    en = _parse_region_edition(_get_xlsx(url_en), {r.nsi_en: r.code for r in REGIONS}, url_en)
    bg = _parse_region_edition(_get_xlsx(url_bg), {r.nsi_bg: r.code for r in REGIONS}, url_bg)

    # The two editions are joined BY LABEL rather than by position, so a
    # reordered sheet cannot mis-pair a name with a figure the way it could in
    # the by-sector table. This check is therefore about something else: the two
    # files must be the same table. НСИ publish identical figures in both, so a
    # code whose series differ means one edition has been revised and the other
    # not, and a payload built from the pair would date half of itself wrongly.
    for r in REGIONS:
        if en["series"][r.code] != bg["series"][r.code]:
            raise ValueError(
                f"The two editions disagree for {r.code!r} ({r.nsi_en} / "
                f"{r.nsi_bg}). НСИ publish the same figures in both files, so "
                f"one of them has been revised without the other. Refusing to "
                f"publish a label from one file against a figure from the other."
            )

    # The headline quarter is the latest one EVERY област carries. A quarter
    # where only some are published would leave the picker with regions that
    # render blank, and a payload whose `ref_period` is true of part of itself.
    common = set.intersection(*(set(en["series"][r.code]) for r in REGIONS))
    ref_period = _latest_published_quarter(dict.fromkeys(common, 0.0))
    if ref_period is None:
        raise ValueError(
            f"No quarter is published for all {len(REGIONS)} области in "
            f"{url_en}, so there is no period the whole table describes."
        )

    regions = [
        {
            "code": r.code,
            "en_name": r.nsi_en.lstrip(_DISTRICT_PREFIX).strip(),
            "bg_name": r.nsi_bg.lstrip(_DISTRICT_PREFIX).strip(),
            "value_eur": en["series"][r.code][ref_period],
            "series_by_period": dict(en["series"][r.code]),
        }
        for r in REGIONS
    ]

    # The third part of the regression guard — see the module docstring for why
    # the other two are the row set rather than a value comparison. Sofia-city
    # is the highest-paid област in Bulgaria by a margin no revision closes
    # (1915 against a next-highest 1304 at 2026-Q1), so a selector that slipped
    # onto the statistical-region headings, or by one row, lands here.
    top = max(regions, key=lambda r: r["value_eur"])
    if top["code"] != SOFIA_CITY_CODE:
        raise ValueError(
            f"Regression guard: the highest wage at {ref_period} is "
            f"{top['code']!r} ({top['value_eur']}), not {SOFIA_CITY_CODE!r}. "
            f"Sofia-city has the highest regional wage in BG, so this means the "
            f"rows were read against the wrong labels."
        )

    return {
        "ref_period": ref_period,
        "is_preliminary": en["prelim_by_period"].get(ref_period, False),
        "regions": regions,
        # НСИ publish no explicit as_of on the workbook; the CLI stamps the
        # payload from the pipeline's own date.today().
        "fetched_at": None,
    }


# ---------------------------------------------------------------------------
# THE BY-SECTOR TABLE — Labour_1.1.2.1, both language editions
# ---------------------------------------------------------------------------

SECTOR_SOURCE_URL_EN = (
    "https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.1_EUR_EN.xlsx"
)
# The Bulgarian edition of the same table carries no language suffix.
SECTOR_SOURCE_URL_BG = (
    "https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.1_EUR.xlsx"
)

# One sheet per year in each edition, named for the classification in that
# edition's language: "2026NaceRev2" and "2026КИД2008". Both files also carry a
# combined "2019-2026…" sheet; the year regex below excludes it by anchoring on
# exactly four digits, which is also why the suffix is matched rather than
# stripped — "2019-2026Кид2008 " differs from its per-year siblings in case and
# in trailing whitespace, and would survive a `.replace()`.
_SECTOR_SHEET_RE_EN = re.compile(r"^(\d{4})NaceRev2$", re.IGNORECASE)
_SECTOR_SHEET_RE_BG = re.compile(r"^(\d{4})КИД2008$", re.IGNORECASE)

# Column 0 of the row that opens a block, per edition. The same string opens the
# MONTHLY block and the QUARTERLY one, so it never identifies a block by itself —
# `_activity_block_header` also requires the period marker below in column 1.
_ACTIVITY_LABEL_EN = "Economic activity"
_ACTIVITY_LABEL_BG = "Икономически дейности"

# Column 1 of that row: "Months" / "Месеци" opens the monthly block, "Quarters
# 2026" / "Тримесечия на 2026 година" the quarterly one. Matched as a prefix
# because the quarterly marker carries the year and the monthly one does not.
_QUARTERLY_MARKER_EN = "quarters"
_QUARTERLY_MARKER_BG = "тримесечия"

# 19 NACE Rev 2 sections plus the all-activities row that heads them.
_SECTOR_ROW_COUNT = 20


def _activity_block_header(rows: list, activity_label: str, quarterly_marker: str) -> int:
    """Index of the row opening the QUARTERLY by-activity block.

    Both the monthly and the quarterly block open with the same column-0 label,
    so the period marker in column 1 is what tells them apart. Returning the row
    INDEX rather than the row is the point: every read below is bounded from it,
    which is what stops a lookup falling through into the ownership block or
    into one of the two `Total` title rows that carry no data.
    """
    for i, row in enumerate(rows):
        if not row or row[0] is None or str(row[0]).strip() != activity_label:
            continue
        marker = str(row[1] or "").strip().lower() if len(row) > 1 else ""
        if marker.startswith(quarterly_marker):
            return i
    raise ValueError(
        f"No quarterly by-activity block found: no row has {activity_label!r} in "
        f"column 0 and a column-1 marker starting {quarterly_marker!r}. НСИ may "
        f"have renamed the block header or dropped the quarterly table."
    )


def _sector_rows(rows: list, header_row: int) -> list[list]:
    """The data rows of one block: from two below its header to the first blank.

    The blank label is the block terminator НСИ's own layout provides. Without
    it a read runs on into the `Type of ownership` block below, whose rows carry
    wages of the same magnitude and would be silently accepted as sectors.
    """
    out: list[list] = []
    for row in rows[header_row + 2 :]:
        if not row or row[0] is None or not str(row[0]).strip():
            break
        out.append(list(row))
    return out


def _parse_sector_sheet(ws, sheet_re, activity_label: str, quarterly_marker: str) -> dict[str, Any]:
    """One per-year sheet -> `{"year", "is_preliminary", "sectors": [(name, {q: eur})]}`.

    Sectors keep НСИ's row ORDER, because that order is what pairs the two
    language editions against each other in `fetch_sector_salary_eu`.
    """
    match = sheet_re.match(str(ws.title).strip())
    if match is None:
        raise ValueError(f"Sheet {ws.title!r} is not a per-year sector sheet.")
    year = int(match.group(1))

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_row = _activity_block_header(rows, activity_label, quarterly_marker)
    cols = _quarter_columns(rows[header_row + 1])
    data = _sector_rows(rows, header_row)
    if len(data) != _SECTOR_ROW_COUNT:
        raise ValueError(
            f"Sheet {ws.title!r} quarterly block has {len(data)} activity rows, "
            f"expected {_SECTOR_ROW_COUNT} (19 NACE Rev 2 sections plus the "
            f"all-activities total). НСИ may have changed the classification."
        )

    # The title carries the preliminary marker for the whole year, as in the
    # regional workbook: "... IN 2026*", "... ПРЕЗ 2026 ГОДИНА*".
    is_preliminary = str(rows[1][0] or "").rstrip().endswith("*")

    sectors = []
    for row in data:
        values = {
            q: float(row[col])
            for col, q in cols.items()
            if col < len(row) and isinstance(row[col], (int, float))
        }
        sectors.append((str(row[0]).strip(), values))
    return {"year": year, "is_preliminary": is_preliminary, "sectors": sectors}


def _parse_sector_edition(raw: bytes, sheet_re, activity_label: str, quarterly_marker: str) -> dict:
    """One language edition -> row-ordered names plus a period-keyed series each."""
    import io

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sheets = [n for n in wb.sheetnames if sheet_re.match(str(n).strip())]
    if not sheets:
        raise ValueError(
            f"No per-year sector sheet matching {sheet_re.pattern} found. "
            f"Available sheets: {wb.sheetnames}"
        )

    names: list[str] = []
    series: list[dict[str, float]] = []
    prelim_by_period: dict[str, bool] = {}
    for name in sorted(sheets):
        parsed = _parse_sector_sheet(wb[name], sheet_re, activity_label, quarterly_marker)
        year = parsed["year"]
        sheet_names = [n for n, _ in parsed["sectors"]]
        if not names:
            names = sheet_names
            series = [{} for _ in names]
        elif sheet_names != names:
            raise ValueError(
                f"Sheet {name!r} lists activities in a different order or under "
                f"different names than the earlier sheets. Pairing the two "
                f"language editions by position is only safe while the order is "
                f"stable, so this is a hard failure rather than a re-sort."
            )
        for i, (_, values) in enumerate(parsed["sectors"]):
            for q, v in values.items():
                series[i][f"{year}-Q{q}"] = v
                prelim_by_period[f"{year}-Q{q}"] = parsed["is_preliminary"]
    return {"names": names, "series": series, "prelim_by_period": prelim_by_period}


def fetch_sector_salary_eu(
    url_en: str = SECTOR_SOURCE_URL_EN,
    url_bg: str = SECTOR_SOURCE_URL_BG,
) -> dict[str, Any]:
    """Fetch НСИ's by-sector average wage table in both language editions.

    Returns:
        {
            "ref_period": "YYYY-Qn",        # latest quarter EVERY sector carries
            "is_preliminary": bool,
            "sectors": [
                {"en_name": str, "bg_name": str, "value_eur": float,
                 "series_by_period": {"YYYY-Qn": float, ...}},
                ...
            ],
        }

    Every value is a cell НСИ published, in the quarter they published it for.
    Nothing here is averaged, rebased, re-levelled or compared — the comparison
    a reader sees is computed in their browser (docs/legal.md §НСИ).

    Raises:
        httpx.HTTPError — network failure
        ValueError      — structure changed, the two editions disagree, or a
                          regression guard tripped
    """
    en = _parse_sector_edition(
        _get_xlsx(url_en), _SECTOR_SHEET_RE_EN, _ACTIVITY_LABEL_EN, _QUARTERLY_MARKER_EN
    )
    bg = _parse_sector_edition(
        _get_xlsx(url_bg), _SECTOR_SHEET_RE_BG, _ACTIVITY_LABEL_BG, _QUARTERLY_MARKER_BG
    )

    if len(en["names"]) != len(bg["names"]):
        raise ValueError(
            f"The two editions list different activity counts: "
            f"{len(en['names'])} English against {len(bg['names'])} Bulgarian."
        )

    # The editions are paired BY POSITION, so this is the check that makes the
    # pairing safe rather than merely convenient: НСИ publish the same figures in
    # both files, so every paired cell must be equal. An inserted or reordered
    # row shifts one edition against the other and lands here, instead of
    # shipping one section's wage under another section's name.
    for i, (en_name, bg_name) in enumerate(zip(en["names"], bg["names"], strict=True)):
        if en["series"][i] != bg["series"][i]:
            raise ValueError(
                f"Row {i} pairs {en_name!r} with {bg_name!r} but their series "
                f"differ, so the two editions are no longer in the same order. "
                f"Refusing to publish a label from one file against a figure "
                f"from the other."
            )

    # The headline quarter is the latest one EVERY sector carries. A quarter
    # where only some are published would leave the picker with sectors that
    # render blank, and a payload whose `ref_period` is true of part of itself.
    common = set.intersection(*(set(s) for s in en["series"])) if en["series"] else set()
    ref_period = _latest_published_quarter(dict.fromkeys(common, 0.0))
    if ref_period is None:
        raise ValueError(
            f"No quarter is published for all {len(en['names'])} activities in "
            f"{url_en}, so there is no period the whole table describes."
        )

    sectors = [
        {
            "en_name": en_name,
            "bg_name": bg["names"][i],
            "value_eur": en["series"][i][ref_period],
            "series_by_period": dict(en["series"][i]),
        }
        for i, en_name in enumerate(en["names"])
    ]

    # Regression guard, the by-sector counterpart of the Sofia-city one. The
    # all-activities row is an average OVER the sections, so it must sit strictly
    # inside their range. It does not if the row selector drifted onto a title
    # row (two carry the word `Total` and no data), onto the public/private
    # ownership block, or onto the monthly table — where the March bonus spike
    # puts several sections above a quarterly total they are below.
    values = [s["value_eur"] for s in sectors[1:]]
    total = sectors[0]["value_eur"]
    if not values or not (min(values) < total < max(values)):
        raise ValueError(
            f"Regression guard: the all-activities row ({total}) does not sit "
            f"between the lowest ({min(values) if values else 'n/a'}) and highest "
            f"({max(values) if values else 'n/a'}) section at {ref_period}. An "
            f"average over the sections must. The block selector likely matched "
            f"the wrong rows."
        )

    return {
        "ref_period": ref_period,
        "is_preliminary": en["prelim_by_period"].get(ref_period, False),
        "sectors": sectors,
        # НСИ publish no explicit as_of on the workbook; the CLI stamps the
        # payload from the pipeline's own date.today().
        "fetched_at": None,
    }


# ---------------------------------------------------------------------------
# Housing — the price and sales workbooks
# ---------------------------------------------------------------------------

# These live in the SAME directory as `Labour_1.1.2.x` above, so the fetch plan,
# the TLS path and the licence read are already understood. What differs is that
# the filenames are **discovered rather than hardcoded**: НСИ's portal lists each
# workbook from a sub-page of a topic index, and a name guessed from a sibling's
# is how this connector would 404 on a rename that a walk survives. The first
# name guessed while researching this — `HPI_1.3.xls` — 404s where `HPI_1.3.xlsx`
# serves.
NSI_PORTAL = "https://www.nsi.bg"
NSI_TIMESERIES_DIR = "https://www.nsi.bg/sites/default/files/files/data/timeseries"

# The workbook each figure comes from, and the topic index that lists it.
#
# `topic` is the `/statistical-data/{id}` page; its sub-pages carry the actual
# links. Held here rather than discovered from the site root because the root
# lists every subject НСИ publish and walking it would be a crawl rather than a
# fetch — the topic id is the smallest thing that has to be written down, and a
# topic that stops listing its own workbook raises.
HOUSING_WORKBOOKS: dict[str, dict[str, str]] = {
    # National house price index, change on the same quarter a year earlier.
    # **The cross-publisher reconciliation reads this**: it is the same
    # statistic as Eurostat's `prc_hpi_q` RCH_A, reaching us by a second route.
    "HPI_1.3": {"topic": "99", "role": "national price index, y/y"},
    # The six cities over 120,000 people, price index y/y.
    "HPI_2.6": {"topic": "98", "role": "six-city price index, y/y"},
    # The same six cities, number of sales, y/y.
    "HSI_2.4.5": {"topic": "93", "role": "six-city sales count, y/y"},
}

# The three rows every one of these workbooks publishes per geography, keyed by
# НСИ's own code. `H.` is the price index and `N.` the sales count; the suffix
# is the same split in both.
HOUSING_ROW_CODES: dict[str, str] = {
    "H.1.": "total",
    "H.1.1.": "new",
    "H.1.2.": "existing",
    "N.1.": "total",
    "N.1.1.": "new",
    "N.1.2.": "existing",
}

# The six cities, in НСИ's own spelling, and the slug each maps to.
#
# A dated hand-authored table, like `regions.py`'s: НСИ name them in Bulgarian
# in the workbook and nothing in the file carries a code the site could join on.
# Read from `HPI_2.4`'s own footnote 2026-08-12: «По данни на НСИ към 31.12.2022
# г. шестте града с население над 120 000 жители са: София, Пловдив, Варна,
# Бургас, Русе и Стара Загора».
#
# **The city label carries a footnote digit** — «Варна 4» is Варна with НСИ's
# marker glued on — so the label is matched after the marker is stripped rather
# than compared whole. One of the six carries it today and any of them may
# tomorrow.
HOUSING_CITIES: dict[str, str] = {
    "София": "sofiya",
    "Пловдив": "plovdiv",
    "Варна": "varna",
    "Бургас": "burgas",
    "Русе": "ruse",
    "Стара Загора": "stara-zagora",
}

# A year header with НСИ's footnote markers glued on. `20263,5` is 2026 wearing
# two of them, and `2026 3` is the same year with a space instead. A
# `str(y).isdigit()` parse drops the newest quarter silently and reads the one
# before it as the latest — which is a plausible number for a wrong period, the
# worst shape a bug takes here.
_HOUSING_YEAR_RE = re.compile(r"^\s*(\d{4})")

# The quarter numerals carry footnotes too — `І6`, `І 7` — and the labour
# workbooks' `_roman_quarter` matches the numeral exactly, so it returns None
# for those and the column is skipped without a word. This is the same map and
# the same Cyrillic translation with the marker stripped first.
_HOUSING_QUARTER_RE = re.compile(r"^([IVХ]+)", re.IGNORECASE)


def _housing_year(cell: object) -> int | None:
    """The year a header cell names, footnote markers and all."""
    if cell is None:
        return None
    m = _HOUSING_YEAR_RE.match(str(cell))
    return int(m.group(1)) if m else None


def _housing_quarter(cell: object) -> int | None:
    """The quarter a header cell names, footnote markers and all."""
    if cell is None:
        return None
    text = str(cell).strip().translate(_CYRILLIC_TO_LATIN)
    m = _HOUSING_QUARTER_RE.match(text)
    return _ROMAN_TO_QUARTER.get(m.group(1).upper()) if m else None


def _strip_footnote(label: object) -> str:
    """A geography label without НСИ's trailing footnote digit."""
    return re.sub(r"\s*\d+\s*$", "", str(label or "")).strip()


def discover_housing_workbook(stem: str, topic: str, timeout: float = 30.0) -> str:
    """The URL НСИ currently publish `{stem}.xlsx` at, read off their own pages.

    Walks the topic index to its sub-pages and takes the `timeseries/` link that
    names this workbook. **Raises rather than falling back to a constructed
    URL**: a guessed name that happens to 404 is a network error somebody
    investigates, and a guessed name that happens to resolve to a DIFFERENT
    vintage is a wrong number nobody sees.
    """
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        index = client.get(f"{NSI_PORTAL}/statistical-data/{topic}")
        index.raise_for_status()
        subs = sorted(set(re.findall(rf"statistical-data/{topic}/(\d+)", index.text)))
        if not subs:
            raise ValueError(
                f"НСИ topic {topic} lists no sub-pages, so no workbook can be "
                f"discovered from it. The portal's shape has changed."
            )
        wanted = re.compile(rf"(timeseries/{re.escape(stem)}\.xlsx)")
        for sub in subs:
            page = client.get(f"{NSI_PORTAL}/statistical-data/{topic}/{sub}")
            page.raise_for_status()
            hit = wanted.search(page.text)
            if hit:
                return f"{NSI_TIMESERIES_DIR}/{stem}.xlsx"
    raise ValueError(
        f"НСИ topic {topic} no longer lists {stem}.xlsx on any of its "
        f"{len(subs)} sub-pages. It has been renamed, moved or withdrawn — "
        f"find where it went rather than hardcoding a URL past this."
    )


def _parse_housing_sheet(ws, has_geography: bool) -> dict[str, dict[str, dict[str, float]]]:
    """One housing workbook sheet → `{geography: {code: {"YYYY-Qn": value}}}`.

    `has_geography` is False for the national files, whose rows start at the
    НСИ code with no city column in front of them. The national result is keyed
    under one empty geography so both shapes read the same downstream.

    A city label appears once per three-row block and the rows under it inherit
    it, which is how the workbook is laid out and why the label is carried
    forward rather than read per row.
    """
    header_year_row, header_quarter_row = 4, 5
    columns: dict[int, str] = {}
    year: int | None = None
    for col in range(1, ws.max_column + 1):
        year = _housing_year(ws.cell(header_year_row, col).value) or year
        quarter = _housing_quarter(ws.cell(header_quarter_row, col).value)
        if year and quarter:
            columns[col] = f"{year}-Q{quarter}"
    if not columns:
        raise ValueError(
            "housing workbook: no quarter columns found. The year header carries "
            "НСИ's footnote markers glued to the numeral and the quarter header "
            "carries them too, so an exact-match parse finds nothing here."
        )

    out: dict[str, dict[str, dict[str, float]]] = {}
    geography = ""
    code_col = 2 if has_geography else 1
    for row in range(6, ws.max_row + 1):
        if has_geography:
            label = ws.cell(row, 1).value
            if label:
                geography = _strip_footnote(label)
        code = str(ws.cell(row, code_col).value or "").strip()
        field = HOUSING_ROW_CODES.get(code)
        if field is None:
            continue
        series: dict[str, float] = {}
        for col, period in columns.items():
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                # НСИ publish these to one decimal and the workbook stores the
                # float their own subtraction produced — `-19.200000000000003`
                # for a cell printed as `-19.2`. Rounding to the published
                # precision is reading the cell as they publish it; carrying the
                # artefact through would put a figure on the page that appears
                # on no НСИ table.
                series[period] = round(float(value), 1)
        if series:
            out.setdefault(geography, {})[field] = series
    if not out:
        raise ValueError(
            "housing workbook: no rows matched НСИ's own row codes "
            f"({', '.join(sorted(HOUSING_ROW_CODES))}). The sheet's layout has moved."
        )
    return out


def fetch_housing_workbook(stem: str, timeout: float = 60.0) -> dict[str, Any]:
    """One НСИ housing workbook, discovered, downloaded and parsed.

    Returns `{"stem", "url", "role", "sheet", "data"}` where `data` is
    `{geography: {total|new|existing: {"YYYY-Qn": value}}}`.
    """
    spec = HOUSING_WORKBOOKS[stem]
    url = discover_housing_workbook(stem, spec["topic"], timeout=timeout)
    wb = openpyxl.load_workbook(io.BytesIO(_get_xlsx(url, timeout=timeout)), data_only=True)
    # Matched by parsing the name, never by index: these workbooks carry one
    # sheet today and the family they belong to carries dozens.
    sheet = wb.sheetnames[0]
    has_geography = stem != "HPI_1.3"
    return {
        "stem": stem,
        "url": url,
        "role": spec["role"],
        "sheet": sheet,
        "data": _parse_housing_sheet(wb[sheet], has_geography),
    }
