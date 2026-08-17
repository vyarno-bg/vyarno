"""БНБ XLSX connector — the mortgage outstanding stock.

One number: **the average rate every Bulgarian household with a home loan is
currently paying.** Not what a new borrower is quoted (that is ЕЦБ MIR new
business, `ecb.py`) — this is the average across the whole existing €18 bn
book, including loans signed years ago.

    file  : s_ir_loan_oa_hh_bg.xlsx
    sheet : LOAN_OA_HH
    cell  : Жилищни кредити (housing loans) · в евро (EUR) · maturity total

**Do not use `s_ir_loan_oa_rm_hh_bg.xlsx`.** That workbook covers "loans other
than overdraft for the household sector" — every purpose blended, consumer
credit included, with no housing breakdown at all. Two tells: its column volume
is €28.7 bn (all household lending) against housing's €18.2 bn, and the
neighbouring column reads 14.83%, a rate no mortgage carries.

Cross-checked against ЕЦБ MIR `M.BG.B.A22.A.R.A.2250.EUR.O`: 2.6717% vs 2.67%,
0.002 pp apart, because they are the same data — БНБ is the institution that
reports MIR to the ЕЦБ. `mortgage.py` enforces that agreement as a gate, so if
either side moves independently the pipeline stops.

COLUMN DISCOVERY — WHY WE DO NOT HARDCODE AN INDEX
--------------------------------------------------
The header is a 4-row merged bureaucracy: purpose → currency → maturity. A
hardcoded index drifts into a neighbouring cell without anyone noticing, so we
locate the column by *reading the headers* every run and raise if they no
longer say what we expect:

    row 3  col 25  "Обеми в млн. евро"     → rates left of here, volumes right
    row 4  col  9  "Жилищни кредити"        → housing block (rates)
           col 33  "Жилищни кредити"        → housing block (volumes)
    row 5  col  9  "в евро"                 → EUR sub-block
    row 6  col  9  ""                       → the block's own column = TOTAL
           col 10  "до 1 година"            → maturity buckets start here
           col 11  "над 1 до 5 години"
           col 12  "над 5 години"

We take the *total* column (row 6 blank), not a maturity bucket, because the
honest answer to "what does the average mortgage holder pay" is the whole book.

TLS PREREQUISITE
----------------
www.bnb.bg serves a certificate issued by "GeoTrust EV RSA CA G2" but does
**not** send that intermediate, so a default trust store fails with "unable to
get local issuer certificate". Fetch the intermediate from the leaf's AIA URL
and append it to the CA bundle — never disable verification. The two commands
are in `docs/data-sources.md` §"TLS setup".
"""

from __future__ import annotations

import io
import warnings
from datetime import datetime
from typing import Any

import httpx
import openpyxl

# URL contract — cited verbatim in mortgage.json provenance.
SOURCE_URL = (
    "https://www.bnb.bg/bnbweb/groups/public/documents/bnb_download/s_ir_loan_oa_hh_bg.xlsx"
)
SHEET_NAME = "LOAN_OA_HH"

# The second workbook, and the only place the fixed/floating split survives the
# euro changeover: ЕЦБ MIR publishes the four buckets' RATES but no volumes on
# the euro leg (`sources/ecb.py` §FIXATION_KEYS), so the share of new lending
# that floats is БНБ's to give or nobody's.
#
# Its header grammar is the one above — section / purpose / currency / bucket on
# the same four rows — so `_locate_housing_eur_columns` finds the housing EUR
# block here unchanged, and the four buckets are the columns to its right.
FIXATION_URL = (
    "https://www.bnb.bg/bnbweb/groups/public/documents/bnb_download/s_ir_loan_nbf_hh_bg.xlsx"
)
FIXATION_SHEET = "LOAN_NBF_HH"
# The trailing digit is БНБ's own footnote marker, and the footnote is what the
# first bucket's label needs: «Включват се кредитите с променлив лихвен
# процент». It is variable-rate loans plus one-year fixations, so no surface may
# call it «fixed for a year» — and no surface may call the rest «fixed» without
# saying for how long, which is the whole point of publishing four buckets.
FIXATION_LABELS = ("до 1 година2", "над 1 до 5 години", "над 5 до 10 години", "над 10 години")
FIXATION_BUCKETS = ("up_to_1y", "1y_to_5y", "5y_to_10y", "over_10y")

# Header labels we require. If БНБ renames any of these we raise, rather than
# read a neighbouring cell.
LABEL_VOLUMES = "Обеми в млн. евро"  # marks where the volume half starts
LABEL_HOUSING = "Жилищни кредити"  # the purpose block we want
LABEL_EUR = "в евро"  # the currency sub-block we want

# Zero-based row indices of the header rows (openpyxl `values_only` rows).
ROW_SECTION = 2  # row 3: rates | volumes
ROW_PURPOSE = 3  # row 4: consumer | housing | other
ROW_CURRENCY = 4  # row 5: EUR | USD
ROW_MATURITY = 5  # row 6: (blank = total) | <=1y | 1-5y | >5y


def fetch_housing_stock_rate_bg() -> list[dict[str, Any]]:
    """Monthly average rate on the outstanding BG housing-loan stock (EUR).

    Returns one dict per month, oldest first:
        {"period": "2026-05", "rate_pct": 2.6717, "volume_eur_m": 18163.0}

    Raises:
        httpx.HTTPError on network/TLS failure (caller exits 4).
        ValueError if the sheet or headers changed (caller exits 2).
    """
    with httpx.Client(timeout=60.0, follow_redirects=True) as client:
        r = client.get(SOURCE_URL)
        r.raise_for_status()
        body = r.content
    return parse_housing_stock_xlsx(body)


def fetch_housing_fixation_bg() -> list[dict[str, Any]]:
    """New housing lending split by initial rate-fixation period (EUR).

    One dict per month, oldest first, with a value per bucket in
    `FIXATION_BUCKETS`:
        {"period": "2026-06", "total_eur_m": 767.075, "total_rate_pct": 2.4066,
         "volume_eur_m": {"up_to_1y": 763.908, ...},
         "rate_pct": {"up_to_1y": 2.4053, ...}}

    Raises the same two as `fetch_housing_stock_rate_bg`.
    """
    with httpx.Client(timeout=60.0, follow_redirects=True) as client:
        r = client.get(FIXATION_URL)
        r.raise_for_status()
        body = r.content
    return parse_housing_fixation_xlsx(body)


def _locate_housing_eur_columns(rows: list[tuple]) -> tuple[int, int]:
    """Find (rate_col, volume_col) for housing loans in EUR, maturity total.

    Pure header inspection — see the module docstring for the layout. Raises
    ValueError with the actual header contents when anything is off, so the
    CLI's exit-2 message tells the next maintainer what BNB changed.
    """
    if len(rows) <= ROW_MATURITY:
        raise ValueError(
            f"BNB sheet {SHEET_NAME!r} has only {len(rows)} rows; "
            f"expected at least {ROW_MATURITY + 1} header rows."
        )

    def cells(row_idx: int) -> list[str]:
        return [(c.strip() if isinstance(c, str) else "") for c in rows[row_idx]]

    section, purpose, currency, maturity = (
        cells(ROW_SECTION),
        cells(ROW_PURPOSE),
        cells(ROW_CURRENCY),
        cells(ROW_MATURITY),
    )

    # Where does the volume half begin? Everything left of it is a rate.
    volume_start = next((i for i, c in enumerate(section) if c == LABEL_VOLUMES), None)
    if volume_start is None:
        raise ValueError(
            f"BNB: header row {ROW_SECTION + 1} no longer contains "
            f"{LABEL_VOLUMES!r} (the rates/volumes divider). "
            f"Row reads: {[c for c in section if c]!r}"
        )

    housing_cols = [i for i, c in enumerate(purpose) if c == LABEL_HOUSING]
    rate_blocks = [i for i in housing_cols if i < volume_start]
    volume_blocks = [i for i in housing_cols if i >= volume_start]
    if not rate_blocks or not volume_blocks:
        raise ValueError(
            f"BNB: expected {LABEL_HOUSING!r} in both the rates and volumes "
            f"halves of header row {ROW_PURPOSE + 1} (divider at col "
            f"{volume_start}); found it at {housing_cols!r}. "
            f"Row reads: {[(i, c) for i, c in enumerate(purpose) if c]!r}"
        )
    rate_col, volume_col = rate_blocks[0], volume_blocks[0]

    # The housing block must start with the EUR sub-block...
    for col, what in ((rate_col, "rate"), (volume_col, "volume")):
        if currency[col] != LABEL_EUR:
            raise ValueError(
                f"BNB: {what} housing block at col {col} is headed "
                f"{currency[col]!r}, expected {LABEL_EUR!r}. "
                f"BNB may have reordered the currency sub-blocks."
            )
        # ...and its own column must be the maturity TOTAL (blank label),
        # with the buckets starting one column to the right.
        if maturity[col] != "":
            raise ValueError(
                f"BNB: expected a blank maturity label at col {col} "
                f"(the '{LABEL_HOUSING}' total), got {maturity[col]!r}. "
                f"Reading a maturity bucket instead of the total would "
                f"silently change what this number means."
            )
    return rate_col, volume_col


def _sheet_rows(body: bytes, sheet_name: str) -> list[tuple]:
    """One БНБ sheet as rows of values. Shared by both workbooks."""
    # БНБ's workbooks carry a print header openpyxl cannot parse. It is
    # cosmetic and unrelated to the cells we read, so mute just that warning
    # rather than letting it clutter every pipeline run.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer so it will be ignored",
            category=UserWarning,
        )
        try:
            wb = openpyxl.load_workbook(io.BytesIO(body), data_only=True, read_only=True)
        except Exception as e:
            # openpyxl raises BadZipFile / KeyError / others depending on how
            # the body is malformed. They all mean "this is not the workbook we
            # expected", so normalise to ValueError, which the CLI reports as
            # exit 2 — otherwise a БНБ error page served with HTTP 200 surfaces
            # as a bare traceback and exit 1.
            raise ValueError(
                f"BNB response is not a readable XLSX "
                f"({type(e).__name__}: {e}). Got {len(body)} bytes — BNB may "
                f"have served an error page with HTTP 200, or changed the "
                f"download URL."
            ) from e
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not found in BNB XLSX; available sheets: {wb.sheetnames}"
        )
    return list(wb[sheet_name].iter_rows(values_only=True))


def parse_housing_stock_xlsx(body: bytes) -> list[dict[str, Any]]:
    """Parse the BNB workbook → the housing-loan EUR outstanding-rate series.

    Pure function: tests pass the committed fixture, no network.
    """
    rows = _sheet_rows(body, SHEET_NAME)
    rate_col, volume_col = _locate_housing_eur_columns(rows)

    out: list[dict[str, Any]] = []
    for row in rows:
        # Data rows are the ones whose first cell is a real date; header and
        # footnote rows carry strings or blanks.
        if not isinstance(row[0], datetime):
            continue
        rate = row[rate_col] if rate_col < len(row) else None
        if not isinstance(rate, (int, float)):
            raise ValueError(
                f"BNB: housing EUR rate cell (col {rate_col}) is "
                f"{rate!r} for {row[0]:%Y-%m}, expected a number. "
                f"Re-verify the layout against docs/data-sources.md."
            )
        volume = row[volume_col] if volume_col < len(row) else None
        out.append(
            {
                "period": row[0].strftime("%Y-%m"),
                "rate_pct": float(rate),
                "volume_eur_m": (float(volume) if isinstance(volume, (int, float)) else None),
            }
        )
    if not out:
        raise ValueError(
            f"BNB: no dated data rows found in {SHEET_NAME!r}. The workbook layout changed."
        )
    out.sort(key=lambda r: r["period"])
    return out


def parse_housing_fixation_xlsx(body: bytes) -> list[dict[str, Any]]:
    """Parse `s_ir_loan_nbf_hh_bg.xlsx` → new housing lending by fixation.

    Pure function: tests pass the committed fixture, no network.
    """
    rows = _sheet_rows(body, FIXATION_SHEET)
    rate_col, volume_col = _locate_housing_eur_columns(rows)
    labels = [(c.strip() if isinstance(c, str) else "") for c in rows[ROW_MATURITY]]
    # The buckets are the four columns right of each block's total, and their
    # labels are asserted rather than counted on: БНБ ordering the housing block
    # differently would otherwise put «над 10 години» money under «до 1 година»
    # and report a floating market as a fixed one, which is the single claim
    # this workbook exists on the site to make.
    for col in (rate_col, volume_col):
        got = tuple(labels[col + 1 : col + 1 + len(FIXATION_LABELS)])
        if got != FIXATION_LABELS:
            raise ValueError(
                f"BNB: the rate-fixation buckets right of col {col} read {got!r}, "
                f"expected {FIXATION_LABELS!r}. Re-verify the housing block in "
                f"{FIXATION_SHEET!r} before trusting the fixed/floating split."
            )

    def number(row: tuple, col: int) -> float | None:
        cell = row[col] if col < len(row) else None
        return float(cell) if isinstance(cell, (int, float)) else None

    out: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row[0], datetime):
            continue
        out.append(
            {
                "period": row[0].strftime("%Y-%m"),
                "total_eur_m": number(row, volume_col),
                "total_rate_pct": number(row, rate_col),
                "volume_eur_m": {
                    b: number(row, volume_col + 1 + i) for i, b in enumerate(FIXATION_BUCKETS)
                },
                "rate_pct": {
                    b: number(row, rate_col + 1 + i) for i, b in enumerate(FIXATION_BUCKETS)
                },
            }
        )
    if not out:
        raise ValueError(
            f"BNB: no dated data rows found in {FIXATION_SHEET!r}. The workbook layout changed."
        )
    out.sort(key=lambda r: r["period"])
    return out
