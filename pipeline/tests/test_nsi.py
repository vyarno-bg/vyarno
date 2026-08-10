"""Tests for the НСИ (National Statistical Institute) XLSX connectors.

The regional connector reads `Labour_1.1.2.2` in **both** language editions —
one `{year}trimes` sheet per year of НСИ's own published quarterly averages by
statistical region — for every one of the 28 области. The headline quarter is
the latest one they all carry, and each value is the cell at that quarter.
Nothing is averaged here, and `test_no_figure_is_computed_only_selected` is what
holds that: the payload is a straight reproduction, which is what §2.1.1 of
НСИ's licence needs (docs/legal.md §НСИ).

These tests build fixture workbooks that mirror the live sheet structure — title
row with the preliminary marker, quarter header row mixing Cyrillic and Latin
numerals, the trailing "IV incl.annual bonuses" column, the statistical-region
headings the district rows sit under, and the Bulgarian edition's extra «Общо»
row that pushes its headers one row down. A fixture locks the tests to the
schema rather than to НСИ's numbers, which move on their own schedule.
"""

from __future__ import annotations

import io
import re

import openpyxl
import pytest

from vyarno_pipeline import clock
from vyarno_pipeline.regions import REGIONS, SOFIA_CITY_CODE, SOFIA_OBLAST_CODE
from vyarno_pipeline.sources import nsi
from vyarno_pipeline.sources.nsi import (
    SOURCE_URL,
    SOURCE_URL_BG,
    _quarter_columns,
    _roman_quarter,
    fetch_region_salaries_eu,
)
from vyarno_pipeline.transform import build_region_salary_payload, build_sector_salary_payload
from vyarno_pipeline.validate import ValidationError, validate_region_salary

# The live headers, byte for byte: Q1/Q2 are Cyrillic І (U+0406), Q3/Q4 are
# Latin. Copied rather than generated, because a generated header would test
# the generator.
LIVE_QUARTER_HEADERS = ["І", "ІІ", "III", "IV"]
BONUS_HEADER = "IV incl.annual bonuses"
BONUS_HEADER_BG = "IV вкл.годишни премии"

# The statistical-region headings НСИ put above each block of области. They are
# not districts and must never be published as ones — they carry no leading
# hyphen, which is exactly how the parse tells them apart.
_REGION_HEADINGS_EN = ("Severozapaden", "Yugozapaden")
_REGION_HEADINGS_BG = ("Северозападен район", "Югозападен район")

# The fixture's two years, taken from the clock rather than written down.
#
# НСИ publish one sheet per year and the connector picks the latest quarter
# every област carries, so a workbook pinned to literal years stops resembling
# the live one the moment the calendar passes it — and the assertions built on
# it stop testing what they claim. Deriving both years and every expected period
# from `clock.today()` keeps the fixture shaped like this year's workbook for as
# long as this app is up, and means no expectation is a literal that has to be
# found and edited each January.
_LATEST_YEAR = clock.today().year
_PRIOR_YEAR = _LATEST_YEAR - 1
_LATEST_PERIOD = f"{_LATEST_YEAR}-Q1"
_PRIOR_PERIODS = tuple(f"{_PRIOR_YEAR}-Q{q}" for q in (1, 2, 3, 4))

# One plausible wage per област, so the fixture exercises all 28 rows rather
# than a Sofia-shaped subset. Sofia-city is deliberately the maximum, as it is
# live — the third part of the regression guard turns on that ordering.
_BASE_WAGES = {r.code: 950.0 + 10 * i for i, r in enumerate(REGIONS)}
_BASE_WAGES[SOFIA_CITY_CODE] = 1915.0
_BASE_WAGES[SOFIA_OBLAST_CODE] = 1294.0


def _build_edition(
    *,
    bulgarian: bool = False,
    years: dict | None = None,
    quarter_headers: list[str] | None = None,
    wages: dict[str, float] | None = None,
    rename: tuple[str, str] | None = None,
    extra_district: str | None = None,
    drop: str | None = None,
    pad_rows: int = 0,
) -> bytes:
    """One language edition of `Labour_1.1.2.2`, mirroring the live layout.

    `years` maps year -> {"quarters": [1, 2, ...], "is_preliminary": bool} and
    the cell for (code, quarter) is `wages[code] + quarter`, so every published
    figure is distinct and a test can assert which cell was read.

    The Bulgarian edition differs in three ways that the live files differ in,
    and all three are deliberate: sheet names carry surrounding spaces, an extra
    «Общо» row sits above the unit marker so the quarter headers land one row
    lower, and the bonus column is spelled in Bulgarian.
    """
    wages = wages or _BASE_WAGES
    years = years or {
        _PRIOR_YEAR: {"quarters": [1, 2, 3, 4], "is_preliminary": False},
        _LATEST_YEAR: {"quarters": [1], "is_preliminary": True},
    }
    headings = _REGION_HEADINGS_BG if bulgarian else _REGION_HEADINGS_EN
    bonus = BONUS_HEADER_BG if bulgarian else BONUS_HEADER

    wb = openpyxl.Workbook()
    # The combined monthly sheet НСИ ship, present and deliberately empty of
    # anything the connector wants: if the parser ever drifts back to it, these
    # tests fail rather than quietly reading a month.
    wb.active.title = f"2019-{_LATEST_YEAR}"
    wb.active.append(["Статистически райони" if bulgarian else "Statistical regions"])

    for year, spec in sorted(years.items()):
        name = f"{year} trimes " if bulgarian else f"{year}trimes"
        ws = wb.create_sheet(name)
        star = "*" if spec.get("is_preliminary", False) else ""
        ws.append([None])
        for _ in range(pad_rows):
            ws.append([None])
        ws.append(
            [
                (
                    f"СРЕДНА БРУТНА МЕСЕЧНА ЗАПЛАТА ... ПРЕЗ {year} ГОДИНА{star}"
                    if bulgarian
                    else f"AVERAGE GROSS MONTHLY WAGES ... AND DISTRICTS IN {year}{star}"
                )
            ]
        )
        if bulgarian:
            # The extra row the Bulgarian edition carries and the English one
            # does not. It is why the quarter header row is searched for.
            ws.append(["\xa0Общо"])
        ws.append(["(евро)\xa0" if bulgarian else "(EUR) "])
        ws.append(
            [
                "Статистически райони" if bulgarian else "Statistical regions",
                f"Тримесечия на {year} година" if bulgarian else f"Quarters {year}",
            ]
        )
        headers = quarter_headers if quarter_headers is not None else LIVE_QUARTER_HEADERS
        ws.append([None, *headers, bonus])

        ws.append([headings[0], 1117.0, None, None, None, None])
        for r in REGIONS:
            label = r.nsi_bg if bulgarian else r.nsi_en
            if drop is not None and label == drop:
                continue
            if rename and label == rename[0]:
                label = rename[1]
            base = wages[r.code]
            row = [label, *[base + q if q in spec["quarters"] else None for q in (1, 2, 3, 4)]]
            # The bonus column, filled for the quarters that have a Q4. It is
            # ~8% above Q4 and reading it would step the whole ladder.
            row.append(base * 1.08 if 4 in spec["quarters"] else None)
            ws.append(row)
            if r.code == SOFIA_OBLAST_CODE:
                ws.append([headings[1], 1708.0, None, None, None, None])
        if extra_district is not None:
            ws.append([extra_district, 1000.0, None, None, None, None])
        ws.append(["*предварителни данни" if bulgarian else "*Preliminary data"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _serve(monkeypatch: pytest.MonkeyPatch, en: bytes | None = None, bg: bytes | None = None):
    """Make the connector read fixture bytes instead of fetching the live URLs.

    Only the download helper is replaced; every parsing, pairing and
    regression-guard path below it runs exactly as it does in production.
    """
    en = _build_edition() if en is None else en
    bg = _build_edition(bulgarian=True) if bg is None else bg
    monkeypatch.setattr(
        nsi, "_get_xlsx", lambda url, timeout=30.0: bg if url == SOURCE_URL_BG else en
    )


def test_connector_reads_every_oblast_at_the_latest_published_quarter(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The main path: all 28 области, each at НСИ's newest quarterly cell.

    The count is the point. A Sofia-shaped read of this workbook returns one
    plausible number and leaves twenty-seven области unpublished, which is what
    a city picker cannot be built on.
    """
    _serve(monkeypatch)
    result = fetch_region_salaries_eu()

    assert result["ref_period"] == _LATEST_PERIOD
    assert [r["code"] for r in result["regions"]] == [r.code for r in REGIONS]
    by_code = {r["code"]: r for r in result["regions"]}
    assert by_code[SOFIA_CITY_CODE]["value_eur"] == pytest.approx(1916.0)
    assert by_code[SOFIA_OBLAST_CODE]["value_eur"] == pytest.approx(1295.0)
    # Both languages come from НСИ's own row labels, with their layout hyphen
    # stripped and nothing else changed.
    assert by_code[SOFIA_CITY_CODE]["en_name"] == "Sofia cap."
    assert by_code[SOFIA_CITY_CODE]["bg_name"] == "София(столица)"
    assert by_code["veliko-tarnovo"]["bg_name"] == "Велико Търново"
    # The whole published history for one область, both years and nothing else.
    assert set(by_code["varna"]["series_by_period"]) == {*_PRIOR_PERIODS, _LATEST_PERIOD}


def test_the_statistical_region_headings_are_never_read_as_области(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """НСИ stack six statistical-region headings among the 28 district rows.

    They carry wages of the same magnitude — 1117 for Severozapaden against
    995 for Vidin under it — so a parse that took them would publish six extra
    "places" that are not области, at figures that look like wages. The leading
    hyphen on a district row is the only thing that separates the two.
    """
    _serve(monkeypatch)
    result = fetch_region_salaries_eu()

    assert len(result["regions"]) == len(REGIONS) == 28
    assert 1117.0 not in {r["value_eur"] for r in result["regions"]}
    assert 1708.0 not in {r["value_eur"] for r in result["regions"]}
    assert not any(r["en_name"].startswith("-") for r in result["regions"])


def test_no_figure_is_computed_only_selected(monkeypatch: pytest.MonkeyPatch) -> None:
    """Every value returned is a cell НСИ published.

    The property `docs/legal.md` §НСИ turns on, and it is invisible on screen:
    an averaging step reintroduced here would move no number a reader could
    check, so nothing but this would notice. Neither the mean of a область's
    quarters nor the mean across области appears anywhere.
    """
    _serve(monkeypatch)
    result = fetch_region_salaries_eu()

    for r in result["regions"]:
        assert r["value_eur"] == r["series_by_period"][result["ref_period"]]
    values = [r["value_eur"] for r in result["regions"]]
    assert sum(values) / len(values) not in values


def test_the_annual_bonus_column_is_never_read_as_a_quarter(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """НСИ publish Q4 twice and the second column folds in the 13th salary.

    Reading it would step every област up in Q4 and back down in Q1 — an 8%
    swing that no gate downstream would flag, because both figures are plausible
    wages. Both spellings are refused, because both editions are read now and
    the Bulgarian one heads that column «IV вкл.годишни премии».
    """
    _serve(monkeypatch)
    result = fetch_region_salaries_eu()

    varna = next(r for r in result["regions"] if r["code"] == "varna")
    assert varna["series_by_period"][_PRIOR_PERIODS[3]] == pytest.approx(_BASE_WAGES["varna"] + 4)
    assert _roman_quarter(BONUS_HEADER) is None
    assert _roman_quarter(BONUS_HEADER_BG) is None


def test_quarter_headers_parse_in_either_alphabet() -> None:
    """The live headers mix Cyrillic І (U+0406) with Latin I and V.

    Folding one onto the other means the parse survives НСИ normalising the
    encoding in either direction, rather than breaking on a change that is
    invisible in a spreadsheet.
    """
    assert _roman_quarter("І") == 1
    assert _roman_quarter("ІІ") == 2
    assert _roman_quarter(" I ") == 1
    assert _roman_quarter("II") == 2
    assert _roman_quarter("III") == 3
    assert _roman_quarter("iv") == 4
    assert _roman_quarter(None) is None
    assert _roman_quarter("Quarters 2026") is None


def test_all_latin_headers_still_parse(monkeypatch: pytest.MonkeyPatch) -> None:
    """End-to-end version of the above: НСИ fixing their encoding must not
    break the connector."""
    latin = ["I", "II", "III", "IV"]
    _serve(
        monkeypatch,
        en=_build_edition(quarter_headers=latin),
        bg=_build_edition(bulgarian=True, quarter_headers=latin),
    )
    assert fetch_region_salaries_eu()["ref_period"] == _LATEST_PERIOD


def test_a_header_row_without_four_quarters_is_an_error() -> None:
    """Silence is the wrong failure here.

    If НСИ drop a quarter column or rename the bonus column so that it parses
    as one, the connector must stop rather than publish a series with a hole or
    a bonus-inflated Q4 in it.
    """
    with pytest.raises(ValueError, match="four quarter columns"):
        _quarter_columns([None, "І", "ІІ", "III", BONUS_HEADER])
    with pytest.raises(ValueError, match="four quarter columns"):
        _quarter_columns([None, "І", "ІІ", "III", "IV", "IV"])


def test_the_bulgarian_edition_puts_its_headers_one_row_lower(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The two editions are not row-aligned, and the parse may not assume it.

    The Bulgarian file carries an «Общо» row the English one does not, so its
    quarter headers sit at row 5 against row 4. A fixed index right for one file
    reads a blank row in the other and reports "no quarter columns" — a true
    message about the wrong thing. The header row is therefore searched for,
    and this holds that: both editions parse to the same figures.
    """
    _serve(monkeypatch)
    result = fetch_region_salaries_eu()
    assert len(result["regions"]) == 28

    # And the search is bounded rather than open: a sheet whose headers are
    # pushed past the bound fails with a message naming the header row, instead
    # of scanning to the end and reporting a missing область.
    padded = _build_edition(pad_rows=nsi._MAX_HEADER_SEARCH_ROW)
    _serve(monkeypatch, en=padded)
    with pytest.raises(ValueError, match="No quarter header row"):
        fetch_region_salaries_eu()


def test_the_preliminary_marker_comes_from_the_year_title(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """НСИ mark a whole year preliminary with a `*` on the sheet title and drop
    it when the year is final. The flag has to follow the quarter's own year,
    not the newest sheet in the file — and it is read from the first labelled
    row above the headers, which is a different row in the two editions."""
    _serve(monkeypatch)
    assert fetch_region_salaries_eu()["is_preliminary"] is True

    final = {_LATEST_YEAR: {"quarters": [1], "is_preliminary": False}}
    _serve(
        monkeypatch,
        en=_build_edition(years=final),
        bg=_build_edition(bulgarian=True, years=final),
    )
    assert fetch_region_salaries_eu()["is_preliminary"] is False


def test_regression_guard_when_sofia_city_is_not_the_highest_paid(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A mis-selected row leaves every number plausible.

    Sofia-city is the highest-paid област in Bulgaria by a wide margin — 1915
    against a next-highest 1304 at 2026-Q1 — so a read that slipped onto the
    statistical-region headings, or by one row, breaks that ordering while every
    figure it publishes still looks like a wage. This is the part of the guard
    that a `cap > province` comparison could not express, because there is no
    universal "region X beats region Y" between any other pair.
    """
    slipped = {**_BASE_WAGES, SOFIA_CITY_CODE: 1000.0}
    _serve(
        monkeypatch,
        en=_build_edition(wages=slipped),
        bg=_build_edition(bulgarian=True, wages=slipped),
    )
    with pytest.raises(ValueError, match="Regression guard"):
        fetch_region_salaries_eu()


def test_a_renamed_oblast_row_is_an_error(monkeypatch: pytest.MonkeyPatch) -> None:
    """A область НСИ renames must stop the run, not vanish from the picker.

    The failure this catches is silent by construction: a payload missing one of
    the 28 still validates against itself, still renders, and simply offers a
    reader one fewer place to be from.
    """
    renamed = ("-Varna", "-Varna oblast")
    _serve(
        monkeypatch,
        en=_build_edition(rename=renamed),
        bg=_build_edition(bulgarian=True),
    )
    with pytest.raises(ValueError, match=r"does not name|missing"):
        fetch_region_salaries_eu()


def test_an_oblast_row_that_disappears_is_an_error(monkeypatch: pytest.MonkeyPatch) -> None:
    """A row НСИ simply stop publishing, as opposed to one they rename.

    Separate from the rename above because a renamed row trips the
    unknown-district check first, which leaves the missing-set check with no
    case of its own — and a guard nothing exercises is a guard that can be
    deleted without a test going red. Dropping a row reaches it directly.
    """
    _serve(
        monkeypatch,
        en=_build_edition(drop="-Ruse"),
        bg=_build_edition(bulgarian=True),
    )
    with pytest.raises(ValueError, match="is missing 1 of the"):
        fetch_region_salaries_eu()


def test_a_district_row_we_do_not_name_is_an_error(monkeypatch: pytest.MonkeyPatch) -> None:
    """The other direction, and the guard is one-sided without it.

    Checking only that our 28 are present passes a workbook that has grown a
    29th. НСИ splitting or adding an област would then publish 28 of 29 — a
    payload that is complete against our own table and incomplete against
    theirs.
    """
    _serve(
        monkeypatch,
        en=_build_edition(extra_district="-Novo"),
        bg=_build_edition(bulgarian=True, extra_district="-Ново"),
    )
    with pytest.raises(ValueError, match="does not name"):
        fetch_region_salaries_eu()


def test_the_two_regional_editions_must_agree_cell_for_cell(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """НСИ publish identical figures in both files.

    Joining by label rather than by position already stops a name being paired
    with the wrong figure, so this checks something else: that the two files are
    the same vintage. One edition revised and the other not would date half the
    payload wrongly, and both halves would look right on their own.
    """
    revised = {**_BASE_WAGES, "ruse": _BASE_WAGES["ruse"] + 50}
    _serve(
        monkeypatch,
        en=_build_edition(),
        bg=_build_edition(bulgarian=True, wages=revised),
    )
    with pytest.raises(ValueError, match="editions disagree"):
        fetch_region_salaries_eu()


def test_the_headline_quarter_is_one_every_oblast_carries(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A quarter only some области have published is not a period the table
    describes, and it would leave the picker with regions that render blank."""
    _serve(monkeypatch)
    result = fetch_region_salaries_eu()
    assert all(result["ref_period"] in r["series_by_period"] for r in result["regions"])


def test_a_workbook_with_no_quarterly_sheet_is_an_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The monthly sheet alone is not enough. Falling back to it would
    reintroduce the derived figure this connector exists to avoid."""
    wb = openpyxl.Workbook()
    wb.active.title = f"2019-{_LATEST_YEAR}"
    wb.active.append(["Statistical regions"])
    buf = io.BytesIO()
    wb.save(buf)
    _serve(monkeypatch, en=buf.getvalue(), bg=buf.getvalue())

    with pytest.raises(ValueError, match="No per-year quarterly sheet"):
        fetch_region_salaries_eu()


def test_connector_url_is_nsi_timeseries_xlsx() -> None:
    """The connector hits the canonical НСИ timeseries XLSX, in both editions.
    If НСИ move either file, this fails before a mis-extracted number reaches
    production."""
    base = "https://www.nsi.bg/sites/default/files/files/data/timeseries/Labour_1.1.2.2_EUR"
    assert f"{base}_EN.xlsx" == SOURCE_URL
    # The Bulgarian edition carries no language suffix — the same convention the
    # by-sector pair follows, and the reason it cannot be derived by appending.
    assert f"{base}.xlsx" == SOURCE_URL_BG


def test_the_slug_table_covers_every_oblast_exactly_once() -> None:
    """`regions.py` is the join between two publishers, and a duplicate in it
    silently unpairs a wage from a price.

    Exactly one область has no имот.bg page — Софийска област — and that is a
    fact about имот.bg's coverage rather than a row left unfinished, so it is
    asserted rather than tolerated.
    """
    assert len(REGIONS) == 28
    assert len({r.code for r in REGIONS}) == 28
    assert len({r.nsi_en for r in REGIONS}) == 28
    assert len({r.nsi_bg for r in REGIONS}) == 28
    unpriced = [r.code for r in REGIONS if r.imot_slug is None]
    assert unpriced == [SOFIA_OBLAST_CODE]
    assert len({r.imot_slug for r in REGIONS if r.imot_slug}) == 27
    # Every НСИ label is a district row in their own layout, hyphen and all.
    assert all(r.nsi_en.startswith("-") and r.nsi_bg.startswith("-") for r in REGIONS)


def test_transform_builds_the_region_payload() -> None:
    """The transformer shapes the connector dict into the published envelope.

    Pin it, because the SPA joins `region_salary.json` to `city_price.json` by
    `code` and reads `value_eur` / `ref_period` / both names off each row.
    """
    scrape = {
        "ref_period": _LATEST_PERIOD,
        "is_preliminary": True,
        "regions": [
            {
                "code": r.code,
                "en_name": r.nsi_en.lstrip("-").strip(),
                "bg_name": r.nsi_bg.lstrip("-").strip(),
                "value_eur": _BASE_WAGES[r.code],
                "series_by_period": {
                    _PRIOR_PERIODS[3]: 1.0,
                    _LATEST_PERIOD: _BASE_WAGES[r.code],
                },
            }
            for r in REGIONS
        ],
    }
    payload = build_region_salary_payload(
        scrape, as_of=clock.today(), source_url=SOURCE_URL, source_url_bg=SOURCE_URL_BG
    )

    assert payload["payload_name"] == "region_salary"
    assert payload["unit"] == "eur_per_month"
    assert payload["source"] == "nsi"
    assert payload["ref_period"] == _LATEST_PERIOD
    assert payload["is_preliminary"] is True
    assert [r["code"] for r in payload["regions"]] == [r.code for r in REGIONS]
    # Both editions are cited, so a verify link can land on the one the label
    # a reader is looking at came from.
    assert payload["source_url_bg"].endswith("Labour_1.1.2.2_EUR.xlsx")
    # No national figure is derived. The workbook has no Bulgaria row, and an
    # unweighted mean over 28 области is both a производно произведение and
    # wrong arithmetic.
    assert "national" not in {k.lower() for k in payload}
    assert payload["disclaimer"].startswith("An област is НСИ's district, not a city")


def test_the_gate_refuses_a_headline_that_is_not_the_published_cell() -> None:
    """The one property nothing on screen would reveal.

    §2.1.1 of НСИ's licence forbids distributing производни произведения, so a
    headline this pipeline calculated rather than read is a licence breach that
    looks exactly like a correct number.
    """
    codes = [r.code for r in REGIONS]
    good = [
        {
            "code": c,
            "en_name": c,
            "bg_name": c,
            "value_eur": 1000.0,
            "series_by_period": {_LATEST_PERIOD: 1000.0},
        }
        for c in codes
    ]
    validate_region_salary(good, _LATEST_PERIOD, codes)

    derived = [dict(r) for r in good]
    derived[0]["value_eur"] = 1000.5
    with pytest.raises(ValidationError, match="must BE the published cell"):
        validate_region_salary(derived, _LATEST_PERIOD, codes)


def test_the_gate_refuses_a_table_that_is_not_the_full_set_of_codes() -> None:
    """Every consumer joins this file to `city_price.json` by code, so a code
    missing, renamed or duplicated unpairs a wage from a price — and the payload
    would still agree with itself."""
    codes = [r.code for r in REGIONS]
    rows = [
        {
            "code": c,
            "en_name": c,
            "bg_name": c,
            "value_eur": 1000.0,
            "series_by_period": {_LATEST_PERIOD: 1000.0},
        }
        for c in codes
    ]
    with pytest.raises(ValidationError, match=re.escape("not `regions.py#REGIONS`")):
        validate_region_salary(rows[:-1], _LATEST_PERIOD, codes)


def test_the_gate_refuses_a_wage_that_is_not_a_wage() -> None:
    """A column index landing on an index number, a headcount or a percentage
    misses the band by orders of magnitude, and none of the three looks wrong on
    the screen."""
    codes = [r.code for r in REGIONS]
    rows = [
        {
            "code": c,
            "en_name": c,
            "bg_name": c,
            "value_eur": 1000.0,
            "series_by_period": {_LATEST_PERIOD: 1000.0},
        }
        for c in codes
    ]
    rows[3]["series_by_period"] = {_LATEST_PERIOD: 1000.0, _PRIOR_PERIODS[0]: 4.7}
    with pytest.raises(ValidationError, match="not a monthly wage"):
        validate_region_salary(rows, _LATEST_PERIOD, codes)


# ---------------------------------------------------------------------------
# The by-sector table — Labour_1.1.2.1, both language editions
# ---------------------------------------------------------------------------

# The Bulgarian edition heads its bonus column differently, which is why the
# connector carries both markers. Copied from the live file, not translated.
BONUS_HEADER_BG = "IV вкл.годишни премии"

# Nineteen NACE Rev 2 sections in НСИ's own order, plus the all-activities row
# that heads them. Trimmed to the shape that matters — the two `Total` title
# rows, the ownership block and the quarterly/monthly split — with НСИ's own
# missing spaces after commas preserved, because the row labels are matched
# exactly and tidying them is how the lookup stops matching.
_SECTOR_NAMES_EN = [
    "Total",
    "Agriculture,forestry and fishing",
    "Mining and quarrying",
    "Manufacturing",
    "Electricity,gas,steam and air conditioning supply",
    "Water supply,sewerage,waste management and remediation activities",
    "Construction",
    "Wholesale and retail trade;repair of motor vehicles and motorcycles",
    "Transportation and storage",
    "Accommodation and food service activities",
    "Information and communication",
    "Financial and insurance activities",
    "Real estate activities",
    "Professional,scientific and technical activities",
    "Administrative and support service activities",
    "Public administration and defence;compulsory social security",
    "Education",
    "Human health and social work activities",
    "Arts,entertainment and recreation",
    "Other service activities",
]
_SECTOR_NAMES_BG = [
    "Общо",
    "Селско, горско и рибно стопанство",
    "Добивна промишленост",
    "Преработваща промишленост",
    "Производство и разпределение на електрическа и топлинна енергия",
    "Доставяне на води;канализационни услуги,управление на отпадъци",
    "Строителство",
    "Търговия; ремонт на автомобили и мотоциклети",
    "Транспорт, складиране и съобщения",
    "Хотелиерство и ресторантьорство",
    "Създаване и разпространение на информация и творчески продукти; далекосъобщения",
    "Финансови и застрахователни дейности",
    "Операции с недвижими имоти",
    "Професионални дейности и научни изследвания",
    "Административни и спомагателни дейности",
    "Държавно управление",
    "Образование",
    "Хуманно здравеопазване и социална работа",
    "Култура,спорт и развлечения",
    "Други дейности",
]

# The quarterly figures the fixture publishes for the latest Q1. `Total` sits inside
# the range of the sections, which is the connector's regression guard.
_LATEST_Q1 = [
    1407.0,
    981.0,
    1713.0,
    1201.0,
    2060.0,
    1048.0,
    1110.0,
    1220.0,
    1214.0,
    870.0,
    3176.0,
    2226.0,
    1259.0,
    1923.0,
    1209.0,
    1591.0,
    1437.0,
    1382.0,
    1260.0,
    1014.0,
]

# March is the annual bonus peak, so the MONTHLY block carries visibly
# different figures. If the parser ever reads the monthly table instead, the
# assertions below land on these rather than on the published quarter.
_MARCH_BONUS_MULTIPLE = 1.14
_LATEST_MARCH = [v * _MARCH_BONUS_MULTIPLE for v in _LATEST_Q1]


def _build_sector_year_sheet(
    wb: openpyxl.Workbook,
    year: int,
    *,
    names: list[str],
    quarterly: list[float],
    sheet_suffix: str,
    activity_label: str,
    ownership_label: str,
    months_label: str,
    quarters_label: str,
    total_title: str,
    bonus_header: str,
    is_preliminary: bool = True,
    quarter_headers: list[str] | None = None,
) -> None:
    """Append one per-year sheet with all FOUR blocks the live file stacks.

    Monthly by activity, monthly by ownership, quarterly by activity, quarterly
    by ownership — plus the two rows that carry the word `Total` in column 0 and
    no data at all. Those two are the reason a label-only lookup cannot work,
    so a fixture without them would test a sheet НСИ do not publish.
    """
    ws = wb.create_sheet(f"{year}{sheet_suffix}")
    headers = quarter_headers if quarter_headers is not None else LIVE_QUARTER_HEADERS
    star = "*" if is_preliminary else ""

    ws.append([None])
    ws.append([f"AVERAGE GROSS MONTHLY WAGES AND SALARIES ... {year}{star}"])
    ws.append([total_title])  # a TITLE row carrying `Total` and no figures
    ws.append(["(EUR) "])

    # Block 1 — monthly by activity. Twelve month columns; only March is filled,
    # and with the bonus-inflated figure.
    ws.append([activity_label, months_label])
    ws.append([None, *["I", "II", "III", "IV", "V", "VI"]])
    # Derived from this sheet's own quarterly figures rather than from the
    # module constant, so the monthly block has one row per activity whatever
    # list of activities the sheet carries. The bonus multiple is what matters:
    # a parser that reads the monthly table lands on a visibly different number.
    for name, q in zip(names, quarterly, strict=True):
        ws.append([name, None, None, q * _MARCH_BONUS_MULTIPLE])
    ws.append([None])
    ws.append([None])

    # Block 2 — monthly by ownership.
    ws.append(["(EUR) "])
    ws.append([ownership_label, months_label])
    ws.append([None, *["I", "II", "III"]])
    ws.append(["Public sector", None, None, 1412.0])
    ws.append(["Private sector", None, None, 1497.0])
    ws.append([None])
    ws.append([None])

    # Block 3 — quarterly by activity. The one the connector must read.
    ws.append([total_title])  # the SECOND `Total` title row
    ws.append(["(EUR) "])
    ws.append([activity_label, f"{quarters_label} {year}"])
    ws.append([None, *headers, bonus_header])
    for name, q1 in zip(names, quarterly, strict=True):
        # Q4 and the bonus column are filled for the earlier year only; for the
        # current one Q1 is all НСИ have published.
        ws.append([name, q1, None, None, None, None])
    ws.append([None])
    ws.append([None])

    # Block 4 — quarterly by ownership. Immediately after the block above, so a
    # read that does not stop at the blank label runs straight into it.
    ws.append(["(EUR) "])
    ws.append([ownership_label, f"{quarters_label} {year}"])
    ws.append([None, *headers, bonus_header])
    ws.append(["Public sector", 1398.0])
    ws.append(["Private sector", 1410.0])
    ws.append(["*Preliminary data"])


def _sector_edition_spec(bulgarian: bool, names: list[str] | None = None) -> dict:
    """Every label that differs between the two language editions, in one place.

    The tests that give the two YEARS different activity lists or different
    lengths build their sheets one at a time, and they need the same labels this
    does. Writing the pairs out twice is how one copy ends up with «Тримесечия»
    against `Quarters` and a fixture that tests the wrong edition.
    """
    return {
        "names": names or (_SECTOR_NAMES_BG if bulgarian else _SECTOR_NAMES_EN),
        "sheet_suffix": "КИД2008" if bulgarian else "NaceRev2",
        "activity_label": "Икономически дейности" if bulgarian else "Economic activity",
        "ownership_label": "Форма на собственост" if bulgarian else "Type of ownership",
        "months_label": "Месеци" if bulgarian else "Months",
        "quarters_label": "Тримесечия на" if bulgarian else "Quarters",
        "total_title": "\xa0Общо" if bulgarian else "Total",
        "bonus_header": BONUS_HEADER_BG if bulgarian else BONUS_HEADER,
    }


def _sector_workbook(bulgarian: bool) -> openpyxl.Workbook:
    """A workbook with the combined sheet both editions carry, and nothing else.

    That sheet's name must NOT match the per-year pattern; if it ever does, the
    parser reads a sheet whose columns are years rather than quarters.
    """
    wb = openpyxl.Workbook()
    wb.active.title = (
        f"2019-{_LATEST_YEAR}Кид2008 " if bulgarian else f"2019-{_LATEST_YEAR}NaceRev2"
    )
    wb.active.append(["Икономически дейности" if bulgarian else "Economic activity"])
    return wb


def _saved(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_sector_xlsx(
    *,
    bulgarian: bool = False,
    quarterly: list[float] | None = None,
    names: list[str] | None = None,
    **kwargs,
) -> bytes:
    """One language edition of the by-sector workbook, both years."""
    wb = _sector_workbook(bulgarian)
    spec = {**_sector_edition_spec(bulgarian, names), **kwargs}
    _build_sector_year_sheet(wb, _PRIOR_YEAR, quarterly=[v * 0.92 for v in _LATEST_Q1], **spec)
    _build_sector_year_sheet(wb, _LATEST_YEAR, quarterly=quarterly or _LATEST_Q1, **spec)
    return _saved(wb)


def _serve_sector_fixtures(
    monkeypatch: pytest.MonkeyPatch,
    en: bytes | None = None,
    bg: bytes | None = None,
) -> None:
    """Serve each language edition from the URL the connector asks for."""
    en_bytes = en if en is not None else _build_sector_xlsx()
    bg_bytes = bg if bg is not None else _build_sector_xlsx(bulgarian=True)

    def _fake(url: str, timeout: float = 30.0) -> bytes:
        return bg_bytes if url == nsi.SECTOR_SOURCE_URL_BG else en_bytes

    monkeypatch.setattr(nsi, "_get_xlsx", _fake)


def test_sector_connector_reads_the_published_quarter_not_the_month(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The quarterly block, chosen past two `Total` title rows and a monthly one.

    Four blocks share the sheet and two rows carry `Total` with no figures, so a
    label-only lookup finds a row of blanks first and the monthly table second.
    March is the annual bonus peak — 14% above the quarter in this fixture — so
    reading the wrong block overstates every sector without looking wrong.
    """
    _serve_sector_fixtures(monkeypatch)
    result = nsi.fetch_sector_salary_eu()

    assert result["ref_period"] == _LATEST_PERIOD
    assert len(result["sectors"]) == 20
    by_name = {s["en_name"]: s["value_eur"] for s in result["sectors"]}
    assert by_name["Total"] == pytest.approx(1407.0)
    assert by_name["Information and communication"] == pytest.approx(3176.0)
    # And nothing from the monthly block reached the payload.
    assert all(v not in _LATEST_MARCH for v in by_name.values())


def test_the_ownership_block_is_never_read_as_an_activity(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The block ends at the first blank label, not at the end of the sheet.

    `Public sector` and `Private sector` sit two rows below the last activity
    and carry wages of the same magnitude, so a read that runs on accepts them
    as sections and reports 22 activities where НСИ publish 19 plus a total.
    """
    _serve_sector_fixtures(monkeypatch)
    names = {s["en_name"] for s in nsi.fetch_sector_salary_eu()["sectors"]}

    assert "Public sector" not in names
    assert "Private sector" not in names


def test_the_annual_bonus_column_is_never_read_as_a_sector_quarter() -> None:
    """`IV incl.annual bonuses` is a different quantity wearing Q4's name.

    Checked in BOTH spellings, because the by-sector table is read in both
    editions and the Bulgarian one heads that column «IV вкл.годишни премии».
    A guard that fires in one language leaves the other file undefended.
    """
    assert _roman_quarter(BONUS_HEADER) is None
    assert _roman_quarter(BONUS_HEADER_BG) is None
    assert _roman_quarter("IV") == 4

    header = [None, *LIVE_QUARTER_HEADERS, BONUS_HEADER_BG]
    assert _quarter_columns(header) == {1: 1, 2: 2, 3: 3, 4: 4}


def test_the_two_language_editions_must_agree_cell_for_cell(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Positional pairing is only safe while the two editions are in step.

    НСИ publish the same figures in both files, so an inserted or reordered row
    shows up as a value mismatch. Without this the connector would ship section
    J's wage under section K's Bulgarian name — plausible on screen, and wrong
    in exactly the way nobody would query.
    """
    shuffled = list(_LATEST_Q1)
    shuffled[10], shuffled[11] = shuffled[11], shuffled[10]
    _serve_sector_fixtures(monkeypatch, bg=_build_sector_xlsx(bulgarian=True, quarterly=shuffled))

    with pytest.raises(ValueError, match="no longer in the same order"):
        nsi.fetch_sector_salary_eu()


def test_a_year_sheet_that_reorders_the_activities_is_an_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """One edition's own sheets must list the activities in one order.

    The series a sector carries is accumulated ACROSS the per-year sheets by
    POSITION, so a year that reorders its rows welds one section's earlier
    quarters onto another section's later ones — and the headline `value_eur`,
    read at the newest quarter, becomes the wrong section's wage under the right
    section's name.

    The cross-edition check cannot see this. It compares English against
    Bulgarian, and НСИ reorder both files together, so both sides move and every
    paired cell still matches. This is the only guard between a re-cut
    classification and a picker that quietly answers «Строителство» with
    Information and communication's 3176.
    """
    swapped = list(_SECTOR_NAMES_EN)
    swapped[10], swapped[11] = swapped[11], swapped[10]

    def _mixed_edition(bulgarian: bool = False) -> bytes:
        """The prior year in НСИ's order, the latest with two transposed."""
        wb = _sector_workbook(bulgarian)
        ordered = _SECTOR_NAMES_BG if bulgarian else _SECTOR_NAMES_EN
        spec = _sector_edition_spec(bulgarian)
        _build_sector_year_sheet(
            wb, _PRIOR_YEAR, **{**spec, "names": ordered}, quarterly=[v * 0.92 for v in _LATEST_Q1]
        )
        _build_sector_year_sheet(
            wb,
            _LATEST_YEAR,
            **{**spec, "names": ordered if bulgarian else swapped},
            quarterly=_LATEST_Q1,
        )
        return _saved(wb)

    _serve_sector_fixtures(monkeypatch, en=_mixed_edition(), bg=_mixed_edition(bulgarian=True))

    with pytest.raises(ValueError, match="different order or under"):
        nsi.fetch_sector_salary_eu()


def test_an_activity_count_that_is_not_nineteen_plus_the_total_is_an_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """НСИ publish 19 NACE Rev 2 sections and one all-activities row.

    The block is bounded by a blank label, which is layout rather than meaning:
    a stray blank inside the table truncates it, and a filled row where НСИ left
    a gap extends it into whatever sits below. Either way the picker silently
    offers a different list of sections than the classification has, and every
    figure on it still looks like a wage. The count is the only thing that
    notices a section went missing.
    """
    short_values = list(_LATEST_Q1)[:-1]

    def _short_edition(bulgarian: bool = False) -> bytes:
        """Both years one section short — «Други дейности» never published."""
        wb = _sector_workbook(bulgarian)
        spec = _sector_edition_spec(
            bulgarian, list(_SECTOR_NAMES_BG if bulgarian else _SECTOR_NAMES_EN)[:-1]
        )
        _build_sector_year_sheet(
            wb, _PRIOR_YEAR, quarterly=[v * 0.92 for v in short_values], **spec
        )
        _build_sector_year_sheet(wb, _LATEST_YEAR, quarterly=short_values, **spec)
        return _saved(wb)

    _serve_sector_fixtures(monkeypatch, en=_short_edition(), bg=_short_edition(bulgarian=True))

    with pytest.raises(ValueError, match="activity rows, expected 20"):
        nsi.fetch_sector_salary_eu()


def test_the_headline_quarter_is_one_every_activity_carries(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """`ref_period` is the latest quarter published for ALL of them, not for any.

    НСИ fill the quarterly block column by column, and a release caught mid-way
    has a newer quarter for some sections and not others. Taking the latest
    quarter ANY section carries gives the payload a `ref_period` that is true of
    part of itself: the sections that have it render, and the rest are a picker
    entry that answers with nothing.

    Here the latest Q2 exists for exactly one section, so a connector reading the union
    would headline Q2 while eighteen sections stop at Q1.
    """
    _serve_sector_fixtures(monkeypatch)
    ragged = nsi.fetch_sector_salary_eu()
    assert ragged["ref_period"] == _LATEST_PERIOD

    # Give ONE section a further quarter, in both editions so the cell-for-cell
    # check still passes and this guard is the only thing left to trip.
    def _with_extra_quarter(bulgarian: bool = False) -> bytes:
        raw = _build_sector_xlsx(bulgarian=bulgarian)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[f"{_LATEST_YEAR}КИД2008" if bulgarian else f"{_LATEST_YEAR}NaceRev2"]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        header = nsi._activity_block_header(
            rows,
            "Икономически дейности" if bulgarian else "Economic activity",
            "тримесечия" if bulgarian else "quarters",
        )
        # Column C is Q2 in the quarterly block; the row two below the header
        # opens the data and is the all-activities row's first section.
        ws.cell(row=header + 3, column=3, value=1000.0)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    _serve_sector_fixtures(
        monkeypatch, en=_with_extra_quarter(), bg=_with_extra_quarter(bulgarian=True)
    )
    result = nsi.fetch_sector_salary_eu()

    assert result["ref_period"] == _LATEST_PERIOD, (
        "the headline moved to a quarter only one activity carries, so the "
        "payload's ref_period is true of part of itself"
    )
    assert all(result["ref_period"] in s["series_by_period"] for s in result["sectors"]), (
        "an activity has no value at the payload's own ref_period"
    )


def test_the_all_activities_row_must_sit_inside_the_sections(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """An average over the sections cannot be above all of them.

    The by-sector counterpart of the Sofia-city/province guard: if the block
    selector drifts onto a title row or the ownership table, the relationship
    between the total and the sections it averages is what breaks first.
    """
    impossible = [9999.0, *_LATEST_Q1[1:]]
    _serve_sector_fixtures(
        monkeypatch,
        en=_build_sector_xlsx(quarterly=impossible),
        bg=_build_sector_xlsx(bulgarian=True, quarterly=impossible),
    )

    with pytest.raises(ValueError, match="Regression guard"):
        nsi.fetch_sector_salary_eu()


def test_no_sector_figure_is_computed_only_selected(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Every figure returned is a cell НСИ published.

    The same property `test_no_figure_is_computed_only_selected` holds for the
    regional table, and it is invisible on screen for the same reason: an
    averaging step here would move no number a reader could check against the
    workbook. The mean of a sector's two published quarters appears nowhere.
    """
    _serve_sector_fixtures(monkeypatch)
    result = nsi.fetch_sector_salary_eu()

    published = set(_LATEST_Q1) | {v * 0.92 for v in _LATEST_Q1}
    for sector in result["sectors"]:
        assert set(sector["series_by_period"].values()) <= published
        assert sector["value_eur"] == sector["series_by_period"][result["ref_period"]]


def test_sector_labels_are_нси_own_strings_in_both_languages(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The Bulgarian label is НСИ's, never a translation of their English.

    Section J is the case the feature turns on: НСИ call it «Създаване и
    разпространение на информация и творчески продукти; далекосъобщения», which
    no reader mistakes for «ИТ». A translated "Information and communication"
    would invite exactly that reading, so the label has to come from their own
    Bulgarian edition.
    """
    _serve_sector_fixtures(monkeypatch)
    sectors = nsi.fetch_sector_salary_eu()["sectors"]
    j = next(s for s in sectors if s["en_name"] == "Information and communication")

    assert j["bg_name"] == _SECTOR_NAMES_BG[10]
    assert "далекосъобщения" in j["bg_name"]
    assert all(s["bg_name"] and s["en_name"] for s in sectors)


def test_the_sector_preliminary_marker_comes_from_the_year_title(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """НСИ star a whole year on the sheet title until they finalise it.

    The counterpart of `test_the_preliminary_marker_comes_from_the_year_title`
    for the by-sector table, and the flag is load-bearing rather than
    decorative: it is what puts «(предварителни данни)» on the source line
    under the reader's own gap. A parse that stops finding the star fails
    SILENTLY to `False`, and the card then shows a figure НСИ will revise as a
    settled one — more certainty than exists (P4), with no other assertion in
    this file disturbed.
    """
    _serve_sector_fixtures(monkeypatch)
    assert nsi.fetch_sector_salary_eu()["is_preliminary"] is True

    _serve_sector_fixtures(
        monkeypatch,
        en=_build_sector_xlsx(is_preliminary=False),
        bg=_build_sector_xlsx(bulgarian=True, is_preliminary=False),
    )
    assert nsi.fetch_sector_salary_eu()["is_preliminary"] is False


def test_the_transform_carries_the_preliminary_flag_into_the_payload(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The connector reads the star; this is what puts it in the file.

    `test_the_sector_preliminary_marker_comes_from_the_year_title` above stops
    at the scrape. Between it and the reader is `build_sector_salary_payload`,
    and the flag has to survive that step: hardwiring it to `False` there
    publishes a preliminary quarter as settled with every connector assertion
    in this file still passing, because none of them looks at the payload.

    The site suites cannot cover the gap either. They read the COMMITTED
    `sector_salary.json`, so a transform that dropped the flag would look
    correct until a refresh regenerated the file — and then go red in a data
    commit, a long way from the line that caused it.
    """
    _serve_sector_fixtures(monkeypatch)
    scrape = nsi.fetch_sector_salary_eu()
    assert scrape["is_preliminary"] is True

    def _payload(flag: bool) -> dict:
        return build_sector_salary_payload(
            {**scrape, "is_preliminary": flag},
            as_of=clock.today(),
            source_url=nsi.SECTOR_SOURCE_URL_EN,
            source_url_bg=nsi.SECTOR_SOURCE_URL_BG,
        )

    # Both ways, so a hardwired constant fails whichever one it was hardwired
    # to. Asserting only the True case passes a payload stuck at True, which
    # marks every finalised quarter provisional.
    assert _payload(True)["is_preliminary"] is True
    assert _payload(False)["is_preliminary"] is False

    # And the figures reach the payload as the cells they were. The gate holds
    # the same identity from the payload's own side; this holds it across the
    # step that builds it, where a rounding or a unit conversion would go.
    built = _payload(True)
    assert [s["value_eur"] for s in built["sectors"]] == [s["value_eur"] for s in scrape["sectors"]]
    assert built["ref_period"] == scrape["ref_period"]
