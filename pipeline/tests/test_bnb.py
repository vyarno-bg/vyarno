"""Tests for the BNB connector (the mortgage outstanding-stock tier).

The fixture `bnb_housing_loans_oa_hh_bg.xlsx` is the real БНБ workbook — sheet
`LOAN_OA_HH`, household loans broken down by purpose.
We read the Жилищни кредити (housing) × в евро (EUR) × maturity-total cell.

Two classes of test here:

1. **The value is right.** Housing ≈2.67% at 2026-05 on an €18.2 bn book,
   matching ECB MIR's outstanding series to 0.002 pp.

2. **The wrong cell cannot come back.** The predecessor read a different
   workbook (all household loans, no housing breakdown) and published 4.12%
   — consumer credit blended in, overstating the mortgage rate by 1.45 pp.
   The column is now located by reading the headers, and the tests below
   corrupt each header in turn to prove we raise instead of drifting into a
   neighbouring cell.
"""

from pathlib import Path

import httpx
import openpyxl
import pytest
import respx

from vyarno_pipeline.sources.bnb import (
    LABEL_CARD,
    LABEL_EUR,
    LABEL_HOUSING,
    LABEL_OUTSIDE_GRACE,
    LABEL_VOLUMES,
    LOAN_PURPOSES,
    OVERDRAFT_SHEET,
    OVERDRAFT_URL,
    ROW_CURRENCY,
    ROW_MATURITY,
    ROW_PURPOSE,
    ROW_SECTION,
    SHEET_NAME,
    SOURCE_URL,
    _locate_overdraft_card_columns,
    _locate_purpose_eur_columns,
    fetch_housing_stock_rate_bg,
    fetch_loan_stock_bg,
    fetch_overdraft_card_stock_bg,
    parse_housing_stock_xlsx,
    parse_loan_stock_xlsx,
    parse_overdraft_card_stock_xlsx,
)

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "bnb_housing_loans_oa_hh_bg.xlsx"
FIXTURE_BYTES = FIXTURE_PATH.read_bytes()


def fixture_rows() -> list[tuple]:
    wb = openpyxl.load_workbook(FIXTURE_PATH, data_only=True)
    return list(wb[SHEET_NAME].iter_rows(values_only=True))


# ---------------------------------------------------------------------------
# Source contract
# ---------------------------------------------------------------------------


def test_points_at_the_purpose_split_workbook_not_the_blended_one():
    """`s_ir_loan_oa_hh_bg` has a housing column; `..._oa_rm_hh_...` does not.

    The blended workbook has no housing breakdown at all, so its rate mixes
    consumer credit into the mortgage number (+1.45 pp).
    """
    assert SOURCE_URL.endswith("s_ir_loan_oa_hh_bg.xlsx")
    assert "_oa_rm_hh_" not in SOURCE_URL
    assert SHEET_NAME == "LOAN_OA_HH"


def test_fixture_is_a_real_workbook_with_the_expected_sheet():
    assert FIXTURE_PATH.stat().st_size > 10_000, "XLSX too small to be real"
    assert SHEET_NAME in openpyxl.load_workbook(FIXTURE_PATH).sheetnames


def test_workbook_really_does_break_households_down_by_purpose():
    """Guards the premise: the housing label must be present twice.

    Once in the rates half, once in the volumes half.
    """
    purpose_row = [c for c in fixture_rows()[ROW_PURPOSE] if isinstance(c, str)]
    assert purpose_row.count(LABEL_HOUSING) == 2
    assert "Кредити за потребление" in purpose_row, "consumer block missing"


# ---------------------------------------------------------------------------
# Column discovery
# ---------------------------------------------------------------------------


def test_locates_the_housing_eur_total_columns():
    rate_col, volume_col = _locate_purpose_eur_columns(fixture_rows())
    rows = fixture_rows()
    assert rows[ROW_PURPOSE][rate_col] == LABEL_HOUSING
    assert rows[ROW_CURRENCY][rate_col] == LABEL_EUR
    # Blank maturity label == the block total, not a maturity bucket.
    assert (rows[ROW_MATURITY][rate_col] or "") == ""
    # The rate column must sit in the rates half, the volume column after it.
    assert rate_col < volume_col


def test_picks_the_total_column_not_a_maturity_bucket():
    """The buckets sit immediately to the right; we must not grab one."""
    rate_col, _ = _locate_purpose_eur_columns(fixture_rows())
    rows = fixture_rows()
    assert rows[ROW_MATURITY][rate_col + 1] == "до 1 година"
    assert rows[ROW_MATURITY][rate_col + 2] == "над 1 до 5 години"


@pytest.mark.parametrize(
    "row_idx, label, expected_error",
    [
        (ROW_SECTION, LABEL_VOLUMES, "rates/volumes divider"),
        (ROW_PURPOSE, LABEL_HOUSING, "expected 'Жилищни кредити'"),
    ],
)
def test_raises_when_a_required_header_label_disappears(row_idx, label, expected_error):
    """A renamed header must fail loud, never fall through to another cell."""
    rows = fixture_rows()
    rows[row_idx] = tuple(("RENAMED" if c == label else c) for c in rows[row_idx])
    with pytest.raises(ValueError, match=expected_error):
        _locate_purpose_eur_columns(rows)


def test_raises_when_the_currency_sub_blocks_are_reordered():
    """If USD came first we would silently publish a USD rate."""
    rows = fixture_rows()
    rate_col, _ = _locate_purpose_eur_columns(rows)
    rows[ROW_CURRENCY] = tuple(
        ("в щатски долари" if i == rate_col else c) for i, c in enumerate(rows[ROW_CURRENCY])
    )
    with pytest.raises(ValueError, match="expected 'в евро'"):
        _locate_purpose_eur_columns(rows)


def test_raises_when_the_total_column_becomes_a_maturity_bucket():
    """The subtlest drift: still housing, still EUR, but a different concept."""
    rows = fixture_rows()
    rate_col, _ = _locate_purpose_eur_columns(rows)
    rows[ROW_MATURITY] = tuple(
        ("над 5 години" if i == rate_col else c) for i, c in enumerate(rows[ROW_MATURITY])
    )
    with pytest.raises(ValueError, match="expected a blank maturity label"):
        _locate_purpose_eur_columns(rows)


def test_raises_on_a_truncated_sheet():
    with pytest.raises(ValueError, match="expected at least"):
        _locate_purpose_eur_columns([(1,), (2,)])


# ---------------------------------------------------------------------------
# Parsed values
# ---------------------------------------------------------------------------


def test_parses_the_recorded_series():
    rows = parse_housing_stock_xlsx(FIXTURE_BYTES)
    assert len(rows) > 200, "the workbook goes back to 2007-01"
    assert rows[0]["period"] == "2007-01"
    assert rows == sorted(rows, key=lambda r: r["period"])
    assert all(set(r) == {"period", "rate_pct", "volume_eur_m"} for r in rows)


def test_latest_value_matches_the_ecb_outstanding_series():
    """≈2.67% at 2026-05 — the number the cross-check gate compares."""
    latest = parse_housing_stock_xlsx(FIXTURE_BYTES)[-1]
    assert latest["period"] == "2026-05"
    assert latest["rate_pct"] == pytest.approx(2.6717, abs=0.0001)
    assert latest["rate_pct"] == pytest.approx(2.67, abs=0.01)


def test_book_size_is_the_housing_book_not_all_household_lending():
    """~€18 bn is housing. ~€29 bn would mean we're back on the blended file."""
    latest = parse_housing_stock_xlsx(FIXTURE_BYTES)[-1]
    assert latest["volume_eur_m"] == pytest.approx(18_163, abs=200)
    assert latest["volume_eur_m"] < 25_000, (
        "book too large to be housing only — likely reading all household loans"
    )


def test_no_value_is_in_consumer_credit_territory():
    """The blended workbook's cells read 11–15%. None of ours may.

    This is the single clearest signal that we are on the housing column.
    """
    rows = parse_housing_stock_xlsx(FIXTURE_BYTES)
    worst = max(rows, key=lambda r: r["rate_pct"])
    assert worst["rate_pct"] < 10.0, (
        f"{worst['period']} reads {worst['rate_pct']}% — consumer-credit "
        f"territory, so this is not the housing column"
    )


def test_recent_rates_are_in_the_plausible_bg_mortgage_band():
    rows = [r for r in parse_housing_stock_xlsx(FIXTURE_BYTES) if r["period"] >= "2024-01"]
    assert rows
    for r in rows:
        assert 2.0 <= r["rate_pct"] <= 6.0, f"{r['period']}: {r['rate_pct']}%"


def test_raises_when_the_rate_cell_is_not_numeric():
    """A text cell where a rate belongs means the layout moved."""
    wb = openpyxl.load_workbook(FIXTURE_PATH)
    ws = wb[SHEET_NAME]
    rate_col, _ = _locate_purpose_eur_columns(list(ws.iter_rows(values_only=True)))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if hasattr(row[0].value, "strftime"):
            row[rate_col].value = "n/a"
            break
    import io

    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="expected a number"):
        parse_housing_stock_xlsx(buf.getvalue())


# ---------------------------------------------------------------------------
# HTTP layer
# ---------------------------------------------------------------------------


@respx.mock
def test_fetch_downloads_and_parses():
    route = respx.get(SOURCE_URL).mock(return_value=httpx.Response(200, content=FIXTURE_BYTES))
    rows = fetch_housing_stock_rate_bg()
    assert route.called
    assert rows[-1]["rate_pct"] == pytest.approx(2.6717, abs=0.0001)


def test_a_non_xlsx_body_raises_value_error_not_a_bare_zip_error():
    """BNB sometimes serves an error page with HTTP 200.

    Normalising to ValueError is what makes the CLI report exit 2 instead of
    a bare traceback and exit 1.
    """
    with pytest.raises(ValueError, match="not a readable XLSX"):
        parse_housing_stock_xlsx(b"<html>Service unavailable</html>")


@respx.mock
def test_fetch_raises_on_http_error():
    respx.get(SOURCE_URL).mock(return_value=httpx.Response(503, text="down"))
    with pytest.raises(httpx.HTTPStatusError):
        fetch_housing_stock_rate_bg()


@respx.mock
def test_fetch_raises_on_tls_or_network_failure():
    """BNB omits an intermediate cert; a bad trust store must not pass."""
    respx.get(SOURCE_URL).mock(
        side_effect=httpx.ConnectError("unable to get local issuer certificate")
    )
    with pytest.raises(httpx.ConnectError):
        fetch_housing_stock_rate_bg()


# ---------------------------------------------------------------------------
# The other two purposes, and the revolving workbook
# ---------------------------------------------------------------------------
# What the mortgage panel needed was one cell. `credit.json` needs the same
# sheet's other two purposes and a second workbook, and the tests below are
# about the two ways that goes wrong quietly: a purpose's block located by
# counting rather than by reading its own label, and БНБ's «в т.ч.» nesting
# read as three amounts to add up.

OVERDRAFT_FIXTURE = Path(__file__).parent / "fixtures" / "bnb_overdraft_card_oa_hh_bg.xlsx"
OVERDRAFT_BYTES = OVERDRAFT_FIXTURE.read_bytes()


def test_each_purpose_lands_on_its_own_block():
    """Three purposes, three column pairs, none of them shared."""
    rows = fixture_rows()
    located = {p: _locate_purpose_eur_columns(rows, label) for p, label in LOAN_PURPOSES.items()}
    assert len(set(located.values())) == 3, f"two purposes resolved to the same columns: {located}"
    for purpose, (rate_col, volume_col) in located.items():
        assert rows[ROW_PURPOSE][rate_col] == LOAN_PURPOSES[purpose]
        assert rows[ROW_PURPOSE][volume_col] == LOAN_PURPOSES[purpose]
        assert rows[ROW_CURRENCY][rate_col] == LABEL_EUR
        assert (rows[ROW_MATURITY][rate_col] or "") == ""


def test_the_purposes_are_the_sizes_that_tell_them_apart():
    """Consumer ~€11 bn, housing ~€18 bn, «Други кредити» ~€0.3 bn.

    Reading one purpose's column where another belongs is the failure a band
    cannot catch on its own, because every one of these is a believable
    household total. The order of magnitude is what separates them.
    """
    latest = {p: parse_loan_stock_xlsx(FIXTURE_BYTES, p)[-1] for p in LOAN_PURPOSES}
    assert latest["housing"]["volume_eur_m"] > latest["consumer"]["volume_eur_m"]
    assert latest["consumer"]["volume_eur_m"] > 20 * latest["other"]["volume_eur_m"]
    # And the rates separate them the other way round: unsecured consumer
    # credit costs well over twice a secured home loan.
    assert latest["consumer"]["rate_pct"] > 2 * latest["housing"]["rate_pct"]


def test_the_housing_wrapper_is_the_general_parse():
    assert parse_loan_stock_xlsx(FIXTURE_BYTES, "housing") == parse_housing_stock_xlsx(
        FIXTURE_BYTES
    )


def test_the_revolving_workbook_reads_all_three_nested_blocks():
    rows = parse_overdraft_card_stock_xlsx(OVERDRAFT_BYTES)
    assert rows[0]["period"] == "2000-03", "the workbook reaches back to 2000"
    assert rows == sorted(rows, key=lambda r: r["period"])
    latest = rows[-1]
    # БНБ nest these «в т.ч.»: the card balance is part of the card block, and
    # the card block is part of the overdraft one. Summed rather than contained
    # they would report roughly twice what is actually owed.
    assert latest["card_outside_grace_eur_m"] <= latest["card_eur_m"] <= latest["overdraft_eur_m"]
    # The balance being charged interest costs more than the block that
    # contains it, which is the whole reason the sub-column exists.
    assert latest["card_outside_grace_rate_pct"] > latest["card_rate_pct"]


def test_an_uncomputed_cell_comes_back_as_none_rather_than_zero():
    """БНБ write «nc» through the early 2000s, and 0% is a rate.

    Rendered as «0,00%» an `nc` reads as a bank lending for nothing, and as a
    volume it reads as nobody owing anything.
    """
    early = parse_overdraft_card_stock_xlsx(OVERDRAFT_BYTES)[0]
    assert early["card_eur_m"] is None
    assert early["card_rate_pct"] is None


@pytest.mark.parametrize(
    "row_idx, label, expected_error",
    [
        (ROW_PURPOSE, LABEL_CARD, "in both halves"),
        (ROW_MATURITY, LABEL_OUTSIDE_GRACE, "right of the"),
    ],
)
def test_the_revolving_walk_raises_when_a_label_moves(row_idx, label, expected_error):
    """The nesting is asserted from the labels, never counted off."""
    wb = openpyxl.load_workbook(OVERDRAFT_FIXTURE)
    rows = list(wb[OVERDRAFT_SHEET].iter_rows(values_only=True))
    rows[row_idx] = tuple(("RENAMED" if c == label else c) for c in rows[row_idx])
    with pytest.raises(ValueError, match=expected_error):
        _locate_overdraft_card_columns(rows)


def test_the_revolving_walk_raises_when_the_sub_column_becomes_the_total():
    """The subtlest drift in this workbook, and the one that halves a figure.

    БНБ label the block's own column with nothing and put «в т.ч. извън
    безлихвен гратисен период» beside it. If that label ever moved onto the
    block's own column we would read €371 m — the balance being charged
    interest — as €490 m, every card balance including the ones inside the
    interest-free period. Both are card debt and neither looks wrong.
    """
    wb = openpyxl.load_workbook(OVERDRAFT_FIXTURE)
    rows = list(wb[OVERDRAFT_SHEET].iter_rows(values_only=True))
    card_rate_col = _locate_overdraft_card_columns(rows)["card"][0]
    rows[ROW_MATURITY] = tuple(
        (LABEL_OUTSIDE_GRACE if i == card_rate_col else c) for i, c in enumerate(rows[ROW_MATURITY])
    )
    with pytest.raises(ValueError, match="expected a blank label"):
        _locate_overdraft_card_columns(rows)


@respx.mock
def test_fetch_loan_stock_returns_every_purpose_from_one_download():
    route = respx.get(SOURCE_URL).mock(return_value=httpx.Response(200, content=FIXTURE_BYTES))
    stock = fetch_loan_stock_bg()
    assert route.call_count == 1, "three purposes, one request"
    assert set(stock) == set(LOAN_PURPOSES)


@respx.mock
def test_fetch_overdraft_card_points_at_the_revolving_workbook():
    respx.get(OVERDRAFT_URL).mock(return_value=httpx.Response(200, content=OVERDRAFT_BYTES))
    assert OVERDRAFT_URL.endswith("s_ir_ovdr_cc_oa_hh_bg.xlsx")
    assert fetch_overdraft_card_stock_bg()[-1]["overdraft_eur_m"] > 0
