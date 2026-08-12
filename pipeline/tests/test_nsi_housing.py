"""НСИ's housing workbooks — the parser, and the four traps that fail silently.

Against the real workbooks, committed. Every one of the traps below produces a
**plausible number for the wrong thing** rather than an error, which is why they
are tested individually rather than covered by "the payload came out non-empty":

- a year header carrying НСИ's footnote markers glued to the numeral drops the
  newest quarter and reports the one before it as the latest;
- a quarter numeral carrying them does the same to one column;
- a city label carrying one fails the join for that city alone;
- the workbook stores the float НСИ's own subtraction produced, so a cell
  printed as -19.2 reaches the page as -19.200000000000003.

The figures asserted are read from the fixture rather than written down, except
where a literal IS the point — the trap tests name the quarter and the city
because "the newest quarter parses" is the whole claim.
"""

from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from vyarno_pipeline.sources.nsi import (
    HOUSING_CITIES,
    _housing_quarter,
    _housing_year,
    _parse_housing_sheet,
    _strip_footnote,
)
from vyarno_pipeline.transform import build_nsi_housing_payload
from vyarno_pipeline.validate import (
    ValidationError,
    validate_hpi_across_publishers,
    validate_nsi_housing,
)

FIXTURES = Path(__file__).parent / "fixtures"

WORKBOOKS = {
    "HPI_1.3": ("nsi_HPI_1.3.xlsx", False, "national price index, y/y"),
    "HPI_2.6": ("nsi_HPI_2.6.xlsx", True, "six-city price index, y/y"),
    "HSI_2.4.5": ("nsi_HSI_2.4.5.xlsx", True, "six-city sales count, y/y"),
}


def _sheet(name: str):
    wb = openpyxl.load_workbook(io.BytesIO((FIXTURES / name).read_bytes()), data_only=True)
    return wb[wb.sheetnames[0]]


def _parsed(stem: str):
    name, has_geography, _ = WORKBOOKS[stem]
    return _parse_housing_sheet(_sheet(name), has_geography)


@pytest.fixture
def workbooks() -> dict:
    return {
        stem: {
            "stem": stem,
            "url": f"https://www.nsi.bg/sites/default/files/files/data/timeseries/{stem}.xlsx",
            "role": role,
            "sheet": "sheet",
            "data": _parsed(stem),
        }
        for stem, (_, _, role) in WORKBOOKS.items()
    }


# --- the four traps --------------------------------------------------------


@pytest.mark.parametrize(
    ("cell", "expected"),
    [
        ("2015", 2015),
        # НСИ glue their footnote markers straight onto the year. Both forms are
        # in the committed workbooks: `20263,5` carries two markers with no
        # separator, `2026 3` one after a space, `2023 ` a trailing space alone.
        ("20263,5", 2026),
        ("2026 3", 2026),
        ("2023 ", 2023),
        (None, None),
        ("Градове", None),
    ],
)
def test_a_year_header_parses_through_its_footnote_markers(cell, expected):
    """`str(y).isdigit()` returns False for `20263,5`, and that is the whole bug.

    The column is skipped, the quarter before it becomes the latest, and every
    figure on the page is a real НСИ number for a period three months stale.
    Nothing downstream can tell.
    """
    assert _housing_year(cell) == expected


@pytest.mark.parametrize(
    ("cell", "expected"),
    [
        # Cyrillic І (U+0406) for Q1 and Q2, Latin I/V for Q3 and Q4 — one
        # header row, two alphabets.
        ("І", 1),
        (" IІ", 2),
        (" IIІ", 3),
        (" IV", 4),
        # …and the same markers again, on the quarter this time.
        (" І6", 1),
        (" І 7", 1),
        (" І ", 1),
        (None, None),
    ],
)
def test_a_quarter_header_parses_through_alphabet_and_footnote(cell, expected):
    """The labour workbooks' `_roman_quarter` matches the numeral exactly.

    Reused as-is here it returns None for `І6` and the column is dropped without
    a word — which is the newest quarter, in the file whose newest quarter is
    the only one the page shows.
    """
    assert _housing_quarter(cell) == expected


def test_a_city_label_parses_through_its_footnote_digit():
    """«Варна 4» is Варна wearing НСИ's marker.

    One of the six carries it today and any of them may tomorrow, so the join
    strips rather than compares whole. Unstripped, the payload loses exactly one
    city and the other five look correct.
    """
    assert _strip_footnote("Варна 4") == "Варна"
    assert _strip_footnote("Стара Загора") == "Стара Загора"
    assert _strip_footnote(None) == ""
    # And the live workbook still needs it, or this guard is guarding a fixture.
    assert set(_parsed("HPI_2.6")) == set(HOUSING_CITIES)


def test_every_published_value_is_rounded_to_the_precision_nsi_print():
    """The workbook holds the float their own subtraction produced.

    `-19.200000000000003` for a cell printed `-19.2`. Publishing the artefact
    puts a figure on the page that appears on no НСИ table — which is both wrong
    and, under §2.1.1, a figure we produced rather than selected.
    """
    for stem in WORKBOOKS:
        for geography, fields in _parsed(stem).items():
            for field, series in fields.items():
                for period, value in series.items():
                    assert value == round(value, 1), (
                        f"{stem} {geography or 'national'} {field} {period} is "
                        f"{value!r}, which is not a figure НСИ print"
                    )


# --- the parse as a whole --------------------------------------------------


def test_the_newest_quarter_survives_the_parse():
    """The one assertion the four traps above all defeat, from the other side.

    Each of them drops a column or a row and leaves the rest correct, so the
    check that catches all four at once is that the last column carrying a
    number is the last column the parser produced.
    """
    for stem, (name, _has_geography, _) in WORKBOOKS.items():
        ws = _sheet(name)
        last_numeric_col = max(
            col
            for col in range(1, ws.max_column + 1)
            for row in range(6, ws.max_row + 1)
            if isinstance(ws.cell(row, col).value, (int, float))
        )
        year, quarter = None, _housing_quarter(ws.cell(5, last_numeric_col).value)
        for col in range(last_numeric_col, 0, -1):
            year = _housing_year(ws.cell(4, col).value)
            if year:
                break
        assert year and quarter, f"{stem}: the last populated column has no period"
        expected = f"{year}-Q{quarter}"
        parsed = _parsed(stem)
        for geography, fields in parsed.items():
            assert expected in fields["total"], (
                f"{stem} {geography or 'national'} is missing {expected}, the "
                f"last quarter the workbook actually carries"
            )


def test_the_six_cities_are_the_six_the_workbook_names():
    price, deals = _parsed("HPI_2.6"), _parsed("HSI_2.4.5")
    assert set(price) == set(HOUSING_CITIES)
    assert set(deals) == set(HOUSING_CITIES)


def test_a_sheet_with_no_recognisable_period_raises():
    """An empty parse must not become an empty payload.

    If НСИ move the header rows, every column falls out and the natural failure
    is a payload with no figures in it — which passes every gate over its
    CONTENTS, because there are none. Built rather than blanked out of a
    fixture: the real workbooks merge their header cells, and openpyxl refuses
    to write to a merged one.
    """
    blank = openpyxl.Workbook().active
    blank["A6"], blank["B6"] = "София", "H.1."
    blank["D6"] = 12.3
    with pytest.raises(ValueError, match="no quarter columns"):
        _parse_housing_sheet(blank, has_geography=True)


def test_a_sheet_whose_row_codes_moved_raises():
    """The other half of the same failure: periods parse, no row matches.

    НСИ key each row by their own code — `H.1.` for the total price index,
    `N.1.` for the sales count — and a layout change that shifts the code column
    leaves a sheet full of numbers none of which the parser claims.
    """
    moved = openpyxl.Workbook().active
    moved["D4"], moved["D5"] = "2026", " І"
    moved["A6"], moved["B6"] = "София", "NOT.A.CODE"
    moved["D6"] = 12.3
    with pytest.raises(ValueError, match="row codes"):
        _parse_housing_sheet(moved, has_geography=True)


# --- the payload and its gates ---------------------------------------------


def test_the_payload_publishes_only_cells_nsi_published(workbooks):
    payload = build_nsi_housing_payload(workbooks, date(2026, 8, 12))
    validate_nsi_housing(payload)
    for key in ("city_price_index_yoy", "city_deals_yoy"):
        for city in payload[key]["cities"]:
            assert city["value_pct"] == city["series_by_period"][city["ref_period"]]


def test_a_headline_that_is_not_the_published_cell_is_refused(workbooks):
    """§2.1.1 forbids distributing производни произведения.

    A headline this pipeline calculated rather than selected is a licence breach
    that looks exactly like a correct number, and the identity is the only thing
    in the payload that can tell them apart.
    """
    payload = build_nsi_housing_payload(workbooks, date(2026, 8, 12))
    payload["city_price_index_yoy"]["cities"][0]["value_pct"] += 0.1
    with pytest.raises(ValidationError, match="headlines"):
        validate_nsi_housing(payload)


def test_the_two_publishers_house_price_index_must_agree(workbooks):
    """The same statistic by two routes, so equality rather than a band.

    Both mutations below produce a figure that looks entirely reasonable: the
    first reads the wrong purchase type, the second the wrong quarter.
    """
    payload = build_nsi_housing_payload(workbooks, date(2026, 8, 12))
    nsi = payload["national_price_index_yoy"]["series_by_period"]
    agreeing = {"price_index": {"annual_rate_pct": {p: dict(v) for p, v in nsi.items()}}}
    validate_hpi_across_publishers(payload, agreeing)

    swapped = {
        "price_index": {
            "annual_rate_pct": {
                p: {**v, "new": v.get("existing"), "existing": v.get("new")} for p, v in nsi.items()
            }
        }
    }
    with pytest.raises(ValidationError, match="same statistic"):
        validate_hpi_across_publishers(payload, swapped)

    periods = sorted(nsi)
    shifted = {"price_index": {"annual_rate_pct": {p: dict(nsi[p]) for p in periods}}}
    shifted["price_index"]["annual_rate_pct"][periods[-1]] = dict(nsi[periods[-2]])
    with pytest.raises(ValidationError, match="same statistic"):
        validate_hpi_across_publishers(payload, shifted)


def test_two_series_that_share_no_quarter_are_refused(workbooks):
    """Not a tolerance question — non-overlapping windows on one series.

    The gate compares at the newest SHARED quarter so a refresh landing between
    the two publishers' releases reconciles on the previous one instead of
    failing. That leniency has a floor: no overlap at all means one of the two
    is not the series it is taken for.
    """
    payload = build_nsi_housing_payload(workbooks, date(2026, 8, 12))
    with pytest.raises(ValidationError, match="share no quarter"):
        validate_hpi_across_publishers(
            payload, {"price_index": {"annual_rate_pct": {"1999-Q1": {"total": 1.0}}}}
        )
